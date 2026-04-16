import csv
import datetime
import json
import math
import os
import re
import random
import time
import threading
import uuid
import ipaddress
from io import BytesIO

os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"

import cv2
import numpy as np
import pytesseract
import qrcode
from PIL import Image
from paddleocr import PaddleOCR

from django.db import transaction
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.core.files import File
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.db.models import Count
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.csrf import csrf_exempt
from django.core.mail import send_mail
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError

from .models import (
    Attendance,
    Employee,
    Event,
    PassportAttendance,
    PassportVisitor,
    Visitor,
    VisitorAttendance,
)

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"


NGROK_BASE_URL = "https://exospherical-kimberlie-unrefulgent.ngrok-free.dev"

PADDLE_OCR = PaddleOCR(
    use_doc_orientation_classify=False,
    use_doc_unwarping=False,
    use_textline_orientation=False,
    lang='en'
)

PADDLE_OCR_LOCK = threading.Lock()

# Keep this map for display only
COUNTRY_CODE_MAP = {
    "MYS": "Malaysia",
    "JPN": "Japan",
    "IDN": "Indonesia",
    "SGP": "Singapore",
    "THA": "Thailand",
    "BRN": "Brunei",
    "PHL": "Philippines",
    "VNM": "Vietnam",
    "CHN": "China",
    "KOR": "South Korea",
    "PRK": "North Korea",
    "IND": "India",
    "PAK": "Pakistan",
    "BGD": "Bangladesh",
    "LKA": "Sri Lanka",
    "NPL": "Nepal",
    "MMR": "Myanmar",
    "KHM": "Cambodia",
    "LAO": "Laos",
    "USA": "United States",
    "CAN": "Canada",
    "GBR": "United Kingdom",
    "AUS": "Australia",
    "NZL": "New Zealand",
    "DEU": "Germany",
    "FRA": "France",
    "ITA": "Italy",
    "ESP": "Spain",
    "PRT": "Portugal",
    "NLD": "Netherlands",
    "BEL": "Belgium",
    "CHE": "Switzerland",
    "AUT": "Austria",
    "SWE": "Sweden",
    "NOR": "Norway",
    "DNK": "Denmark",
    "FIN": "Finland",
    "IRL": "Ireland",
    "POL": "Poland",
    "CZE": "Czech Republic",
    "TUR": "Turkey",
    "SAU": "Saudi Arabia",
    "ARE": "United Arab Emirates",
    "QAT": "Qatar",
    "KWT": "Kuwait",
    "OMN": "Oman",
    "BHR": "Bahrain",
    "EGY": "Egypt",
    "ZAF": "South Africa",
    "BRA": "Brazil",
    "MEX": "Mexico",
    "RUS": "Russia",
}

# For global passports: keep validation generic
GENERIC_PASSPORT_PATTERN = r"^[A-Z0-9]{6,12}$"


def country_code_to_name(code):
    code = (code or "").strip().upper()
    return COUNTRY_CODE_MAP.get(code, code if code else "Unknown")


def validate_passport_number_by_country(passport_number, country_code_or_name):
    passport_number = re.sub(r'[^A-Z0-9]', '', (passport_number or '').upper())

    if not passport_number:
        return False, "Passport number cannot be empty"

    if not re.match(GENERIC_PASSPORT_PATTERN, passport_number):
        return False, "Passport number format is invalid"

    return True, ""

def fix_common_ocr_errors(text, mode="general"):
    if not text:
        return ""

    text = text.strip().upper()

    if mode == "mrz":
        text = text.replace(" ", "")
        text = text.replace("«", "<").replace("‹", "<").replace("〈", "<")
        text = text.replace("|", "I")

        # keep only MRZ-safe chars
        text = re.sub(r'[^A-Z0-9<]', '', text)

        # normalize repeated noise
        text = text.replace("KKKK", "<<<<")
        text = text.replace("KKK", "<<<")
        text = text.replace("KK", "<<")

        # common OCR fixes inside MRZ
        text = text.replace("5PN", "JPN")
        text = text.replace("2PN", "JPN")
        text = text.replace("7PN", "JPN")
        text = text.replace("25PN", "JPN")
        text = text.replace("MYS:", "MYS")
        text = text.replace("MYSMA", "MYSMA")
        text = text.replace("P<MYSMA", "P<MYSMA")
        text = text.replace("P<JPNNA", "P<JPNNA")
        text = text.replace("O<", "0<")
        text = text.replace("<K<", "<<")

        return text

    if mode == "passport":
        text = re.sub(r'[^A-Z0-9]', '', text)
        if not text:
            return ""

        # OCR often confuses these in passport number
        text = (
            text.replace("O", "0")
                .replace("Q", "0")
                .replace("I", "1")
                .replace("L", "1")
                .replace("S", "5")
                .replace("B", "8")
        )
        return text

    if mode == "name":
        text = text.replace("0", "O")
        text = text.replace("1", "I")
        text = text.replace("5", "S")
        text = re.sub(r'[^A-Z< ]', '', text)
        return text

    return text

def normalize_display_date(value):
    if not value:
        return ""

    value = str(value).strip()
    if not value:
        return ""

    # yyyy-mm-dd -> dd/mm/yyyy
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", value)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"

    # dd-mm-yyyy -> dd/mm/yyyy
    m = re.match(r"^(\d{2})-(\d{2})-(\d{4})$", value)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"

    # mm/dd/yyyy -> dd/mm/yyyy
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", value)
    if m:
        first = int(m.group(1))
        second = int(m.group(2))

        if first <= 12 and second <= 31:
            return f"{m.group(2)}/{m.group(1)}/{m.group(3)}"

    return value.replace("-", "/")


def is_reasonable_name(value):
    if not value:
        return False

    value = str(value).strip()
    if len(value) < 2:
        return False

    if re.search(r'\d', value):
        return False

    banned = [
        'passport', 'nationality', 'country', 'date', 'issue',
        'expiry', 'birth', 'sex', 'male', 'female', 'identity'
    ]
    lower_value = value.lower()
    if any(word in lower_value for word in banned):
        return False

    cleaned = re.sub(r"[^A-Za-z\s'/-]", "", value).strip()
    return len(cleaned) >= 2


def clean_person_name(value):
    if not value:
        return ""

    value = value.replace("<", " ")
    value = re.sub(r"[^A-Za-z\s'/-]", " ", value)
    value = re.sub(r"\s+", " ", value).strip()

    if not is_reasonable_name(value):
        return ""

    return value

def rescue_mrz_lines(text):
    if not text:
        return []

    raw_lines = [x.strip().upper() for x in text.splitlines() if x.strip()]
    rescued = []

    for line in raw_lines:
        line = fix_common_ocr_errors(line, mode="mrz")
        line = re.sub(r'[^A-Z0-9<]', '', line)

        # passport MRZ lines usually long and contain <
        if len(line) >= 25 and '<' in line:
            rescued.append(line)

    return rescued


def parse_mrz_rescue(text):
    mrz_lines = rescue_mrz_lines(text)
    if len(mrz_lines) < 2:
        return None

    best_pair = None
    best_score = -1

    for i in range(len(mrz_lines) - 1):
        l1 = mrz_lines[i]
        l2 = mrz_lines[i + 1]

        score = 0
        if l1.startswith('P<'):
            score += 2
        if len(l1) >= 35:
            score += 1
        if len(l2) >= 35:
            score += 1
        if '<' in l1 and '<' in l2:
            score += 1

        if score > best_score:
            best_score = score
            best_pair = (l1, l2)

    if not best_pair:
        return None

    return parse_two_line_passport_mrz(best_pair[0], best_pair[1], rescue_mode=True)

def get_current_employee(request):
    if not request.user.is_authenticated:
        return None
    return Employee.objects.filter(user=request.user).first()


def is_admin_user(request):
    employee = get_current_employee(request)
    return bool(employee and employee.role == 'admin')


def is_editor_user(request):
    employee = get_current_employee(request)
    return bool(employee and employee.role == 'editor')


def can_manage_user(request):
    employee = get_current_employee(request)
    return bool(employee and employee.role in ['admin', 'editor'])


def role_context(request):
    current_employee = get_current_employee(request)
    return {
        'current_employee': current_employee,
        'is_admin': bool(current_employee and current_employee.role == 'admin'),
        'is_editor': bool(current_employee and current_employee.role == 'editor'),
        'is_viewer': bool(current_employee and current_employee.role == 'viewer'),
        'can_manage': bool(current_employee and current_employee.role in ['admin', 'editor']),
    }


def require_login_page(request):
    if not request.user.is_authenticated:
        return redirect('/login-page/')
    return None


def require_admin_api(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Login required'}, status=401)
    if not is_admin_user(request):
        return JsonResponse({'error': 'Only admin can perform this action'}, status=403)
    return None


def require_manage_api(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Login required'}, status=401)
    if not can_manage_user(request):
        return JsonResponse({'error': 'You do not have permission for this action'}, status=403)
    return None


def require_manage_page(request):
    if not request.user.is_authenticated:
        return redirect('/login-page/')
    if not can_manage_user(request):
        return HttpResponse('Forbidden', status=403)
    return None


@csrf_exempt
def register_manual(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            full_name = data.get('full_name')
            employee_id = data.get('employee_id')
            email = data.get('email')
            department = data.get('department')
            registration_method = data.get('registration_method', 'manual')
            password = data.get('password')
            confirm_password = data.get('confirm_password')

            if not full_name or not employee_id or not email or not department or not password or not confirm_password:
                return JsonResponse({'error': 'All fields are required'}, status=400)

            if password != confirm_password:
                return JsonResponse({'error': 'Password and confirm password do not match'}, status=400)

            if len(password) < 8:
                return JsonResponse({'error': 'Password must be at least 8 characters'}, status=400)

            if not re.search(r'[A-Za-z]', password) or not re.search(r'\d', password):
                return JsonResponse({'error': 'Password must contain letters and numbers'}, status=400)

            if User.objects.filter(username=employee_id).exists():
                return JsonResponse({'error': 'Employee ID already exists as login user'}, status=400)

            if Employee.objects.filter(employee_id=employee_id).exists():
                return JsonResponse({'error': 'Employee ID already exists'}, status=400)

            if Employee.objects.filter(email=email).exists():
                return JsonResponse({'error': 'Email already exists'}, status=400)

            user = User.objects.create_user(
                username=employee_id,
                email=email,
                password=password
            )

            # CHECK: ada admin sudah wujud atau belum
            if not Employee.objects.filter(role='admin').exists():
                role = 'admin'
            else:
                role = 'viewer'

            Employee.objects.create(
                user=user,
                full_name=full_name,
                employee_id=employee_id,
                email=email,
                department=department,
                registration_method=registration_method,
                role=role
            )

            return JsonResponse({'message': f'Employee registered via {registration_method}'})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request'}, status=400)


def register_page(request):
    return render(request, 'register.html')


@csrf_exempt
def login_user(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            identifier = data.get('username')
            password = data.get('password')

            if not identifier or not password:
                return JsonResponse({'error': 'Username/Email and password required'}, status=400)

            user = None

            if '@' in identifier:
                try:
                    user_obj = User.objects.get(email=identifier)
                    user = authenticate(request, username=user_obj.username, password=password)
                except User.DoesNotExist:
                    return JsonResponse({'error': 'User not found'}, status=404)
            else:
                user = authenticate(request, username=identifier, password=password)

            if user is not None:
                login(request, user)
                return JsonResponse({'message': 'Login successful'})
            else:
                return JsonResponse({'error': 'Invalid credentials'}, status=401)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request'}, status=400)


def login_page(request):
    if request.user.is_authenticated:
        return redirect('/dashboard/')
    return render(request, 'login.html')


def logout_user(request):
    logout(request)
    return redirect('/login-page/')

def clear_forgot_password_session(request):
    keys = [
        'forgot_password_email',
        'forgot_password_user_id',
        'forgot_password_otp',
        'forgot_password_otp_expires_at',
        'forgot_password_verified',
    ]
    for key in keys:
        request.session.pop(key, None)

def forgot_password_page(request):
    if request.user.is_authenticated:
        return redirect('/dashboard/')
    return render(request, 'forgot_password.html')


@csrf_exempt
def send_forgot_password_otp(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        data = json.loads(request.body)
        email = (data.get('email') or '').strip().lower()

        if not email:
            return JsonResponse({'error': 'Email is required'}, status=400)

        try:
            user = User.objects.get(email__iexact=email)
        except User.DoesNotExist:
            return JsonResponse({'error': 'No account found with this email'}, status=404)

        otp = generate_otp()

        request.session['forgot_password_email'] = email
        request.session['forgot_password_user_id'] = user.id
        request.session['forgot_password_otp'] = otp
        request.session['forgot_password_otp_expires_at'] = int(time.time()) + 300  # 5 minit
        request.session['forgot_password_verified'] = False
        request.session.modified = True

        send_mail(
            subject='DBKU Attendance Password Reset OTP',
            message=(
                f'Your OTP code is: {otp}\n\n'
                f'This code will expire in 5 minutes.\n'
                f'If you did not request this, please ignore this email.'
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            fail_silently=False,
        )

        return JsonResponse({'message': 'OTP sent successfully to your email'})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def verify_forgot_password_otp(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        data = json.loads(request.body)
        email = (data.get('email') or '').strip().lower()
        otp = (data.get('otp') or '').strip()

        saved_email = request.session.get('forgot_password_email')
        saved_otp = request.session.get('forgot_password_otp')
        expires_at = request.session.get('forgot_password_otp_expires_at')

        if not saved_email or not saved_otp or not expires_at:
            return JsonResponse({'error': 'Please request OTP first'}, status=400)

        if email != saved_email:
            return JsonResponse({'error': 'Email does not match the OTP request'}, status=400)

        if int(time.time()) > int(expires_at):
            clear_forgot_password_session(request)
            return JsonResponse({'error': 'OTP expired. Please request a new OTP'}, status=400)

        if otp != saved_otp:
            return JsonResponse({'error': 'Invalid OTP code'}, status=400)

        request.session['forgot_password_verified'] = True
        request.session.modified = True

        return JsonResponse({
            'message': 'OTP verified successfully',
            'redirect_url': '/reset-password/'
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def reset_password_page(request):
    if request.user.is_authenticated:
        return redirect('/dashboard/')

    verified = request.session.get('forgot_password_verified')
    user_id = request.session.get('forgot_password_user_id')

    if not verified or not user_id:
        return redirect('/forgot-password/')

    return render(request, 'reset_password.html')


@csrf_exempt
def reset_password_submit(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        verified = request.session.get('forgot_password_verified')
        user_id = request.session.get('forgot_password_user_id')

        if not verified or not user_id:
            return JsonResponse({'error': 'Unauthorized password reset session'}, status=403)

        data = json.loads(request.body)
        new_password = data.get('new_password') or ''
        confirm_password = data.get('confirm_password') or ''

        if not new_password or not confirm_password:
            return JsonResponse({'error': 'All fields are required'}, status=400)

        if new_password != confirm_password:
            return JsonResponse({'error': 'Password and confirm password do not match'}, status=400)

        if len(new_password) < 8:
            return JsonResponse({'error': 'Password must be at least 8 characters'}, status=400)

        if not re.search(r'[A-Za-z]', new_password) or not re.search(r'\d', new_password):
            return JsonResponse({'error': 'Password must contain letters and numbers'}, status=400)

        user = User.objects.get(id=user_id)

        try:
            validate_password(new_password, user=user)
        except ValidationError as e:
            return JsonResponse({'error': ' '.join(e.messages)}, status=400)

        user.set_password(new_password)
        user.save()

        clear_forgot_password_session(request)

        return JsonResponse({
            'message': 'Password reset successful',
            'redirect_url': '/login-page/'
        })

    except User.DoesNotExist:
        clear_forgot_password_session(request)
        return JsonResponse({'error': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def generate_otp():
    return str(random.randint(100000, 999999))

def validate_staff_password_fields(password, confirm_password=None):
    if not password:
        return False, 'Password is required'

    if confirm_password is not None and password != confirm_password:
        return False, 'Password and confirm password do not match'

    if len(password) < 8:
        return False, 'Password must be at least 8 characters'

    if not re.search(r'[A-Za-z]', password) or not re.search(r'\d', password):
        return False, 'Password must contain letters and numbers'

    return True, ''

def dashboard(request):
    login_check = require_login_page(request)
    if login_check:
        return login_check

    today = datetime.date.today()

    total_employees = Employee.objects.count()
    total_visitors = Visitor.objects.count()
    total_passport_visitors = PassportVisitor.objects.count()
    total_events = Event.objects.count()

    active_events_qs = Event.objects.filter(start_date__lte=today, end_date__gte=today).order_by('start_date', 'start_time', 'id')
    upcoming_events_qs = Event.objects.filter(start_date__gt=today).order_by('start_date', 'start_time', 'id')[:5]

    active_events = active_events_qs.count()

    total_staff_attendance = Attendance.objects.count()
    total_visitor_attendance = VisitorAttendance.objects.count()
    total_passport_attendance = PassportAttendance.objects.count()
    total_attendance = total_staff_attendance + total_visitor_attendance + total_passport_attendance

    today_staff_attendance = Attendance.objects.filter(date=today).count()
    today_visitor_attendance = VisitorAttendance.objects.filter(date=today).count()
    today_passport_attendance = PassportAttendance.objects.filter(date=today).count()
    today_total_attendance = today_staff_attendance + today_visitor_attendance + today_passport_attendance

    recent_staff = [
        {
            'name': att.full_name,
            'type': 'Staff',
            'event_name': att.event.name,
            'time': att.time,
            'date': att.date,
        }
        for att in Attendance.objects.select_related('event').order_by('-date', '-time')[:5]
    ]

    recent_visitors = [
        {
            'name': att.visitor.full_name,
            'type': 'Visitor (Malaysian)',
            'event_name': att.event.name,
            'time': att.time,
            'date': att.date,
        }
        for att in VisitorAttendance.objects.select_related('event', 'visitor').order_by('-date', '-time')[:5]
    ]

    recent_passports = [
        {
            'name': att.passport_visitor.full_name,
            'type': 'Visitor (Non-Malaysian)',
            'event_name': att.event.name,
            'time': att.time,
            'date': att.date,
        }
        for att in PassportAttendance.objects.select_related('event', 'passport_visitor').order_by('-date', '-time')[:5]
    ]

    recent_activities = sorted(
        recent_staff + recent_visitors + recent_passports,
        key=lambda x: (x['date'], x['time']),
        reverse=True
    )[:8]

    context = {
        'today_date': today,
        'total_employees': total_employees,
        'total_visitors': total_visitors,
        'total_passport_visitors': total_passport_visitors,
        'total_events': total_events,
        'active_events': active_events,
        'total_staff_attendance': total_staff_attendance,
        'total_visitor_attendance': total_visitor_attendance,
        'total_passport_attendance': total_passport_attendance,
        'total_attendance': total_attendance,
        'today_staff_attendance': today_staff_attendance,
        'today_visitor_attendance': today_visitor_attendance,
        'today_passport_attendance': today_passport_attendance,
        'today_total_attendance': today_total_attendance,
        'active_events_list': active_events_qs[:5],
        'upcoming_events': upcoming_events_qs,
        'recent_activities': recent_activities,
    }
    context.update(role_context(request))

    return render(request, 'dashboard.html', context)


def analytics_page(request):
    login_check = require_login_page(request)
    if login_check:
        return login_check

    event_name = request.GET.get('name', '').strip()
    event_month = request.GET.get('month', '').strip()
    event_year = request.GET.get('year', '').strip()
    event_location = request.GET.get('location', '').strip()

    events = Event.objects.all().order_by('-start_date', '-id')

    if event_name:
        events = events.filter(name__icontains=event_name)

    if event_month:
        events = events.filter(start_date__month=event_month)

    if event_year:
        events = events.filter(start_date__year=event_year)

    if event_location:
        events = events.filter(location__icontains=event_location)

    event_analytics = []

    total_filtered_events = events.count()
    total_filtered_staff = 0
    total_filtered_visitors = 0
    total_filtered_passport = 0

    department_totals = {}
    organization_totals = {}
    country_totals = {}

    monthly_map = {
        'Jan': 0, 'Feb': 0, 'Mar': 0, 'Apr': 0, 'May': 0, 'Jun': 0,
        'Jul': 0, 'Aug': 0, 'Sep': 0, 'Oct': 0, 'Nov': 0, 'Dec': 0
    }

    for event in events:
        staff_group = (
            Attendance.objects.filter(event=event)
            .exclude(department__isnull=True)
            .exclude(department__exact='')
            .values('department')
            .annotate(total=Count('id'))
            .order_by('-total', 'department')
        )

        visitor_group = (
            VisitorAttendance.objects.filter(event=event)
            .select_related('visitor')
            .exclude(visitor__organization__isnull=True)
            .exclude(visitor__organization__exact='')
            .values('visitor__organization')
            .annotate(total=Count('id'))
            .order_by('-total', 'visitor__organization')
        )

        passport_group = (
            PassportAttendance.objects.filter(event=event)
            .select_related('passport_visitor')
            .exclude(passport_visitor__country__isnull=True)
            .exclude(passport_visitor__country__exact='')
            .values('passport_visitor__country')
            .annotate(total=Count('id'))
            .order_by('-total', 'passport_visitor__country')
        )

        staff_labels = [item['department'] for item in staff_group]
        staff_data = [item['total'] for item in staff_group]

        visitor_labels = [item['visitor__organization'] for item in visitor_group]
        visitor_data = [item['total'] for item in visitor_group]

        passport_labels = [item['passport_visitor__country'] for item in passport_group]
        passport_data = [item['total'] for item in passport_group]

        staff_total = sum(staff_data)
        visitor_total = sum(visitor_data)
        passport_total = sum(passport_data)
        grand_total = staff_total + visitor_total + passport_total

        total_filtered_staff += staff_total
        total_filtered_visitors += visitor_total
        total_filtered_passport += passport_total

        if event.start_date:
            month_key = event.start_date.strftime('%b')
            if month_key in monthly_map:
                monthly_map[month_key] += grand_total

        for item in staff_group:
            dept = item['department']
            department_totals[dept] = department_totals.get(dept, 0) + item['total']

        for item in visitor_group:
            org = item['visitor__organization']
            organization_totals[org] = organization_totals.get(org, 0) + item['total']

        for item in passport_group:
            country = item['passport_visitor__country']
            country_totals[country] = country_totals.get(country, 0) + item['total']

        event_analytics.append({
            'event': event,
            'staff_labels_json': json.dumps(staff_labels),
            'staff_data_json': json.dumps(staff_data),
            'visitor_labels_json': json.dumps(visitor_labels),
            'visitor_data_json': json.dumps(visitor_data),
            'passport_labels_json': json.dumps(passport_labels),
            'passport_data_json': json.dumps(passport_data),
            'staff_total': staff_total,
            'visitor_total': visitor_total,
            'passport_total': passport_total,
            'grand_total': grand_total,
        })

    top_departments = sorted(department_totals.items(), key=lambda x: x[1], reverse=True)[:5]
    top_organizations = sorted(organization_totals.items(), key=lambda x: x[1], reverse=True)[:5]
    top_countries = sorted(country_totals.items(), key=lambda x: x[1], reverse=True)[:5]
    top_events = sorted(event_analytics, key=lambda x: x['grand_total'], reverse=True)[:5]

    overview_labels = ['Staff', 'Visitor (Malaysian)', 'Visitor (Non-Malaysian)']
    overview_values = [total_filtered_staff, total_filtered_visitors, total_filtered_passport]

    monthly_labels = list(monthly_map.keys())
    monthly_values = list(monthly_map.values())

    context = {
        'event_analytics': event_analytics,
        'filter_name': event_name,
        'filter_month': event_month,
        'filter_year': event_year,
        'filter_location': event_location,
        'total_filtered_events': total_filtered_events,
        'total_filtered_staff': total_filtered_staff,
        'total_filtered_visitors': total_filtered_visitors,
        'total_filtered_passport': total_filtered_passport,
        'overview_labels_json': json.dumps(overview_labels),
        'overview_values_json': json.dumps(overview_values),
        'monthly_labels_json': json.dumps(monthly_labels),
        'monthly_values_json': json.dumps(monthly_values),
        'top_departments': top_departments,
        'top_organizations': top_organizations,
        'top_countries': top_countries,
        'top_events': top_events,
    }
    context.update(role_context(request))

    return render(request, 'analytics.html', context)


def employees_page(request):
    login_check = require_login_page(request)
    if login_check:
        return login_check

    employees = Employee.objects.all().order_by('full_name')

    search_name = request.GET.get('name')
    department = request.GET.get('department')

    if search_name:
        employees = employees.filter(full_name__icontains=search_name)

    if department:
        employees = employees.filter(department__icontains=department)

    context = {
        'employees': employees
    }
    context.update(role_context(request))

    return render(request, 'employees.html', context)


@csrf_exempt
def add_employee(request):
    permission_check = require_admin_api(request)
    if permission_check:
        return permission_check

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            full_name = (data.get('full_name') or '').strip()
            employee_id = (data.get('employee_id') or '').strip()
            email = (data.get('email') or '').strip().lower()
            department = (data.get('department') or '').strip()
            registration_method = (data.get('registration_method') or 'manual').strip()
            role = (data.get('role') or 'viewer').strip()
            password = data.get('password') or ''
            confirm_password = data.get('confirm_password') or ''

            if not full_name or not employee_id or not email or not department:
                return JsonResponse({'error': 'All fields are required'}, status=400)

            if role not in ['admin', 'editor', 'viewer']:
                return JsonResponse({'error': 'Invalid role'}, status=400)

            is_valid_password, password_error = validate_staff_password_fields(password, confirm_password)
            if not is_valid_password:
                return JsonResponse({'error': password_error}, status=400)

            if User.objects.filter(username=employee_id).exists():
                return JsonResponse({'error': 'Employee ID already exists as login user'}, status=400)

            if Employee.objects.filter(employee_id=employee_id).exists():
                return JsonResponse({'error': 'Employee ID already exists'}, status=400)

            if Employee.objects.filter(email=email).exists():
                return JsonResponse({'error': 'Email already exists'}, status=400)

            user = User.objects.create_user(
                username=employee_id,
                email=email,
                password=password
            )

            Employee.objects.create(
                user=user,
                full_name=full_name,
                employee_id=employee_id,
                email=email,
                department=department,
                registration_method=registration_method,
                role=role
            )

            return JsonResponse({'message': 'Employee added successfully'})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request'}, status=400)

@csrf_exempt
def import_employees_excel(request):
    permission_check = require_admin_api(request)
    if permission_check:
        return permission_check

    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        if 'excel_file' not in request.FILES:
            return JsonResponse({'error': 'Excel file is required'}, status=400)

        excel_file = request.FILES['excel_file']
        filename = excel_file.name.lower()

        if not filename.endswith('.xlsx'):
            return JsonResponse({'error': 'Only .xlsx file is supported'}, status=400)

        workbook = load_workbook(excel_file)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return JsonResponse({'error': 'Excel file is empty'}, status=400)

        headers = [str(cell).strip().lower() if cell is not None else '' for cell in rows[0]]

        required_headers = ['full_name', 'employee_id', 'email', 'department', 'password']
        for header in required_headers:
            if header not in headers:
                return JsonResponse({
                    'error': f'Missing required column: {header}'
                }, status=400)

        header_index = {header: idx for idx, header in enumerate(headers)}

        created_count = 0
        errors = []

        for row_number, row in enumerate(rows[1:], start=2):
            if row is None:
                continue

            full_name = str(row[header_index['full_name']]).strip() if row[header_index['full_name']] is not None else ''
            employee_id = str(row[header_index['employee_id']]).strip() if row[header_index['employee_id']] is not None else ''
            email = str(row[header_index['email']]).strip().lower() if row[header_index['email']] is not None else ''
            department = str(row[header_index['department']]).strip() if row[header_index['department']] is not None else ''
            password = str(row[header_index['password']]).strip() if row[header_index['password']] is not None else ''

            role = 'viewer'
            if 'role' in header_index and row[header_index['role']] is not None:
                role = str(row[header_index['role']]).strip().lower() or 'viewer'

            registration_method = 'manual'
            if 'registration_method' in header_index and row[header_index['registration_method']] is not None:
                registration_method = str(row[header_index['registration_method']]).strip().lower() or 'manual'

            if not full_name and not employee_id and not email and not department and not password:
                continue

            if not full_name or not employee_id or not email or not department or not password:
                errors.append(f'Row {row_number}: Missing required fields')
                continue

            if role not in ['admin', 'editor', 'viewer']:
                errors.append(f'Row {row_number}: Invalid role')
                continue

            is_valid_password, password_error = validate_staff_password_fields(password)
            if not is_valid_password:
                errors.append(f'Row {row_number}: {password_error}')
                continue

            if User.objects.filter(username=employee_id).exists():
                errors.append(f'Row {row_number}: Employee ID already exists as login user')
                continue

            if Employee.objects.filter(employee_id=employee_id).exists():
                errors.append(f'Row {row_number}: Employee ID already exists')
                continue

            if Employee.objects.filter(email=email).exists():
                errors.append(f'Row {row_number}: Email already exists')
                continue

            try:
                with transaction.atomic():
                    user = User.objects.create_user(
                        username=employee_id,
                        email=email,
                        password=password
                    )

                    Employee.objects.create(
                        user=user,
                        full_name=full_name,
                        employee_id=employee_id,
                        email=email,
                        department=department,
                        registration_method=registration_method,
                        role=role
                    )

                created_count += 1

            except Exception as e:
                errors.append(f'Row {row_number}: {str(e)}')

        return JsonResponse({
            'message': f'Excel import completed. {created_count} staff created.',
            'created_count': created_count,
            'error_count': len(errors),
            'errors': errors[:20]
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def export_employees_excel(request):
    permission_check = require_admin_api(request)
    if permission_check:
        return permission_check

    search_name = (request.GET.get('name') or '').strip()
    department = (request.GET.get('department') or '').strip()

    employees = Employee.objects.all().order_by('full_name')

    if search_name:
        employees = employees.filter(full_name__icontains=search_name)

    if department:
        employees = employees.filter(department__icontains=department)

    wb = Workbook()
    ws = wb.active
    ws.title = "Staff"

    headers = [
        'Full Name',
        'Staff ID',
        'Email',
        'Department',
        'Role',
        'Registration Method',
        'Created At',
    ]
    ws.append(headers)

    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for emp in employees:
        ws.append([
            emp.full_name,
            emp.employee_id,
            emp.email,
            emp.department,
            emp.role,
            emp.registration_method,
            emp.created_at.strftime('%d/%m/%Y %H:%M') if emp.created_at else '',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="staff_export.xlsx"'

    wb.save(response)
    return response

@csrf_exempt
def delete_employee(request, id):
    permission_check = require_admin_api(request)
    if permission_check:
        return permission_check

    try:
        emp = Employee.objects.get(id=id)

        if emp.user:
            emp.user.delete()
        else:
            emp.delete()

        return JsonResponse({'message': 'Deleted successfully'})
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def update_employee(request, id):
    permission_check = require_admin_api(request)
    if permission_check:
        return permission_check

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            emp = Employee.objects.get(id=id)

            full_name = (data.get('full_name') or '').strip()
            employee_id = (data.get('employee_id') or '').strip()
            email = (data.get('email') or '').strip().lower()
            department = (data.get('department') or '').strip()
            registration_method = (data.get('registration_method') or '').strip()
            role = (data.get('role') or emp.role).strip()
            password = data.get('password') or ''
            confirm_password = data.get('confirm_password') or ''

            if not full_name or not employee_id or not email or not department:
                return JsonResponse({'error': 'All fields are required'}, status=400)

            if role not in ['admin', 'editor', 'viewer']:
                return JsonResponse({'error': 'Invalid role'}, status=400)

            if Employee.objects.exclude(id=id).filter(employee_id=employee_id).exists():
                return JsonResponse({'error': 'Employee ID already exists'}, status=400)

            if Employee.objects.exclude(id=id).filter(email=email).exists():
                return JsonResponse({'error': 'Email already exists'}, status=400)

            emp.full_name = full_name
            emp.employee_id = employee_id
            emp.email = email
            emp.department = department
            emp.registration_method = registration_method or emp.registration_method
            emp.role = role
            emp.save()

            if emp.user:
                emp.user.username = employee_id
                emp.user.email = email

                if password or confirm_password:
                    is_valid_password, password_error = validate_staff_password_fields(password, confirm_password)
                    if not is_valid_password:
                        return JsonResponse({'error': password_error}, status=400)

                    emp.user.set_password(password)

                emp.user.save()

            return JsonResponse({'message': 'Updated successfully'})

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request'}, status=400)


@csrf_exempt
def create_event(request):
    permission_check = require_manage_api(request)
    if permission_check:
        return permission_check

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            name = (data.get('name') or '').strip()
            location = (data.get('location') or '').strip()
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            start_time = data.get('start_time') or None
            end_time = data.get('end_time') or None

            if not name or not location or not start_date or not end_date:
                return JsonResponse({'error': 'Name, location, start date and end date are required'}, status=400)

            start_date_obj = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date_obj = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()

            if end_date_obj < start_date_obj:
                return JsonResponse({'error': 'End date cannot be earlier than start date'}, status=400)

            if start_date == end_date and start_time and end_time:
                start_time_obj = datetime.datetime.strptime(start_time, "%H:%M").time()
                end_time_obj = datetime.datetime.strptime(end_time, "%H:%M").time()

                if end_time_obj <= start_time_obj:
                    return JsonResponse({'error': 'End time must be later than start time'}, status=400)

            radius = data.get('radius_meter')

            try:
                radius = int(radius)
                if radius <= 0:
                    radius = 100
            except (TypeError, ValueError):
                radius = 100

            event = Event.objects.create(
                name=name,
                location=location,
                start_date=start_date,
                end_date=end_date,
                start_time=start_time,
                end_time=end_time,
                description=data.get('description'),
                latitude=data.get('latitude') if data.get('latitude') not in ['', None] else None,
                longitude=data.get('longitude') if data.get('longitude') not in ['', None] else None,
                radius_meter=radius
            )

            qr_configs = [
                (
                    'visitor_qr_code',
                    f'visitor_event_{event.id}.png',
                    f"{NGROK_BASE_URL}/api/employees/visitor-attendance/{event.id}/",
                ),
                (
                    'staff_qr_code',
                    f'staff_event_{event.id}.png',
                    f"{NGROK_BASE_URL}/api/employees/staff-attendance/{event.id}/",
                ),
                (
                    'passport_qr_code',
                    f'passport_event_{event.id}.png',
                    f"{NGROK_BASE_URL}/api/employees/passport-attendance/{event.id}/",
                ),
            ]

            for field_name, file_name, qr_data in qr_configs:
                qr_image = qrcode.make(qr_data)
                qr_buffer = BytesIO()
                qr_image.save(qr_buffer, format='PNG')
                qr_content = ContentFile(qr_buffer.getvalue())
                getattr(event, field_name).save(file_name, qr_content, save=False)

            event.save()
            event.refresh_from_db()

            return JsonResponse({
                'message': 'Event created successfully',
                'visitor_qr_url': event.visitor_qr_code.url if event.visitor_qr_code else '',
                'staff_qr_url': event.staff_qr_code.url if event.staff_qr_code else '',
                'passport_qr_url': event.passport_qr_code.url if event.passport_qr_code else '',
            })

        except ValueError:
            return JsonResponse({'error': 'Invalid date or time format'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request'}, status=400)


def events_page(request):
    login_check = require_login_page(request)
    if login_check:
        return login_check

    events = Event.objects.all().order_by('-start_date', '-id')

    search_name = request.GET.get('name')
    search_location = request.GET.get('location')
    search_date = request.GET.get('date')

    if search_name:
        events = events.filter(name__icontains=search_name)

    if search_location:
        events = events.filter(location__icontains=search_location)

    if search_date:
        events = events.filter(start_date__lte=search_date, end_date__gte=search_date)

    context = {
        'events': events,
    }
    context.update(role_context(request))

    return render(request, 'events.html', context)

@csrf_exempt
def update_event(request, id):
    permission_check = require_manage_api(request)
    if permission_check:
        return permission_check

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            event = Event.objects.get(id=id)

            new_start_date = data.get('start_date') or str(event.start_date)
            new_end_date = data.get('end_date') or str(event.end_date)
            new_start_time = data.get('start_time') if 'start_time' in data else (event.start_time.strftime("%H:%M") if event.start_time else None)
            new_end_time = data.get('end_time') if 'end_time' in data else (event.end_time.strftime("%H:%M") if event.end_time else None)

            start_date_obj = datetime.datetime.strptime(new_start_date, "%Y-%m-%d").date()
            end_date_obj = datetime.datetime.strptime(new_end_date, "%Y-%m-%d").date()

            if end_date_obj < start_date_obj:
                return JsonResponse({'error': 'End date cannot be earlier than start date'}, status=400)

            if new_start_date == new_end_date and new_start_time and new_end_time:
                start_time_obj = datetime.datetime.strptime(new_start_time, "%H:%M").time()
                end_time_obj = datetime.datetime.strptime(new_end_time, "%H:%M").time()

                if end_time_obj <= start_time_obj:
                    return JsonResponse({'error': 'End time must be later than start time'}, status=400)

            event.name = data.get('name') or event.name
            event.location = data.get('location') or event.location
            event.start_date = new_start_date
            event.end_date = new_end_date
            event.start_time = new_start_time or None
            event.end_time = new_end_time or None
            event.description = data.get('description') if data.get('description') is not None else event.description

            if 'latitude' in data:
                event.latitude = data.get('latitude') if data.get('latitude') not in ['', None] else None

            if 'longitude' in data:
                event.longitude = data.get('longitude') if data.get('longitude') not in ['', None] else None

            if 'radius_meter' in data and data.get('radius_meter') not in ['', None]:
                event.radius_meter = data.get('radius_meter')

            event.save()

            return JsonResponse({'message': 'Updated successfully'})

        except Event.DoesNotExist:
            return JsonResponse({'error': 'Event not found'}, status=404)
        except ValueError:
            return JsonResponse({'error': 'Invalid date or time format'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request'}, status=400)


@csrf_exempt
def delete_event(request, id):
    permission_check = require_manage_api(request)
    if permission_check:
        return permission_check

    try:
        event = Event.objects.get(id=id)
        event.delete()
        return JsonResponse({'message': 'Deleted successfully'})
    except Event.DoesNotExist:
        return JsonResponse({'error': 'Event not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def calculate_distance_meters(lat1, lon1, lat2, lon2):
    lat1 = float(lat1)
    lon1 = float(lon1)
    lat2 = float(lat2)
    lon2 = float(lon2)

    r = 6371000

    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)

    a = math.sin(delta_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    return r * c

def normalize_ip_value(raw_ip):
    if not raw_ip:
        return None

    raw_ip = raw_ip.strip()

    # contoh: [2001:db8::1]:443
    if raw_ip.startswith('[') and ']' in raw_ip:
        raw_ip = raw_ip[1:raw_ip.index(']')]

    # contoh: 192.168.0.10:8000
    elif raw_ip.count(':') == 1 and '.' in raw_ip.split(':')[0]:
        raw_ip = raw_ip.split(':')[0]

    # buang scope id kalau ada, contoh fe80::1%eth0
    raw_ip = raw_ip.split('%')[0]

    try:
        return str(ipaddress.ip_address(raw_ip))
    except ValueError:
        return None


def get_client_ips(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')

    candidates = []
    if x_forwarded_for:
        candidates = [ip.strip() for ip in x_forwarded_for.split(',') if ip.strip()]
    else:
        remote_addr = request.META.get('REMOTE_ADDR')
        if remote_addr:
            candidates = [remote_addr]

    ipv4_address = None
    ipv6_address = None

    for raw_ip in candidates:
        clean_ip = normalize_ip_value(raw_ip)
        if not clean_ip:
            continue

        parsed_ip = ipaddress.ip_address(clean_ip)

        if parsed_ip.version == 4 and not ipv4_address:
            ipv4_address = clean_ip
        elif parsed_ip.version == 6 and not ipv6_address:
            ipv6_address = clean_ip

    return ipv4_address, ipv6_address

def visitor_attendance_page(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    return render(request, 'visitor_attendance_form.html', {'event': event})


@csrf_exempt
def submit_visitor_attendance(request, event_id):
    try:
        if request.method != 'POST':
            return JsonResponse({'error': 'Invalid request'}, status=400)

        data = json.loads(request.body)

        full_name = (data.get('full_name') or '').strip()
        phone = (data.get('phone') or '').strip()
        email = (data.get('email') or '').strip().lower()
        organization = (data.get('organization') or '').strip()
        latitude = data.get('latitude')
        longitude = data.get('longitude')

        if not full_name or not phone or not email or not organization:
            return JsonResponse({'error': 'All fields are required'}, status=400)

        if not phone.isdigit() or len(phone) < 9:
            return JsonResponse({'error': 'Invalid phone number'}, status=400)        

        if latitude in [None, ''] or longitude in [None, '']:
            return JsonResponse({'error': 'Please enable GPS/location first'}, status=400)

        event = get_object_or_404(Event, id=event_id)

        if event.latitude is None or event.longitude is None:
            return JsonResponse({'error': 'Event location not set'}, status=400)

        distance = calculate_distance_meters(latitude, longitude, event.latitude, event.longitude)

        if distance > event.radius_meter:
            return JsonResponse({
                'error': f'Attendance rejected. Outside allowed area ({round(distance, 2)}m)'
            }, status=400)

        visitor, created_visitor = Visitor.objects.get_or_create(
            email=email,
            defaults={
                'full_name': full_name,
                'phone_number': phone,
                'organization': organization
            }
        )

        if not created_visitor:
            visitor.full_name = full_name
            visitor.phone_number = phone
            visitor.organization = organization
            visitor.save()

        attendance, created = VisitorAttendance.objects.get_or_create(
            visitor=visitor,
            event=event,
            defaults={
                'latitude': latitude,
                'longitude': longitude
            }
        )

        if not created:
            return JsonResponse({'error': 'You have already registered for this event'}, status=400)

        return JsonResponse({
            'message': 'Successfully Registered',
            'distance_meter': round(distance, 2),
            'event': event.name
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def staff_attendance_page(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    return render(request, 'staff_attendance_form.html', {'event': event})


@csrf_exempt
def submit_staff_attendance(request, event_id):
    try:
        if request.method != 'POST':
            return JsonResponse({'error': 'Invalid request'}, status=400)

        data = json.loads(request.body)

        full_name = (data.get('full_name') or '').strip()
        employee_id = (data.get('employee_id') or '').strip().upper()
        phone = (data.get('phone') or '').strip()
        email = (data.get('email') or '').strip().lower()
        department = (data.get('department') or '').strip()
        latitude = data.get('latitude')
        longitude = data.get('longitude')
        ipv4_address, ipv6_address = get_client_ips(request)
        submitted_ipv4 = (data.get('ipv4_address') or '').strip()

        # fallback: guna IPv4 dari browser kalau header tak ada
        if not ipv4_address and submitted_ipv4:
            try:
                parsed = ipaddress.ip_address(submitted_ipv4)
                if parsed.version == 4:
                    ipv4_address = str(parsed)
            except ValueError:
                pass        
        
        print("=== STAFF ATTENDANCE IP DEBUG ===")
        print("HTTP_X_FORWARDED_FOR:", request.META.get('HTTP_X_FORWARDED_FOR'))
        print("REMOTE_ADDR:", request.META.get('REMOTE_ADDR'))
        print("IPv4 detected:", ipv4_address)
        print("IPv6 detected:", ipv6_address)

        if not all([full_name, employee_id, phone, email, department]):
            return JsonResponse({'error': 'All fields are required'}, status=400)

        if not phone.isdigit() or len(phone) < 9:
            return JsonResponse({'error': 'Invalid phone number'}, status=400)

        # ✅ GPS CHECK
        if latitude in [None, ''] or longitude in [None, '']:
            return JsonResponse({'error': 'Please enable GPS/location first'}, status=400)

        # ✅ EVENT CHECK (SAFE)
        event = get_object_or_404(Event, id=event_id)

        if event.latitude is None or event.longitude is None:
            return JsonResponse({'error': 'Event location not configured by admin'}, status=400)

        # ✅ VALIDATE EMPLOYEE EXIST
        employee = Employee.objects.filter(employee_id=employee_id).first()

        if not employee:
            return JsonResponse({'error': 'Employee ID not found. Please register first.'}, status=404)

        # OPTIONAL STRICT CHECK (email match)
        if employee.email.lower() != email:
            return JsonResponse({'error': 'Email does not match registered employee'}, status=400)

        # ✅ GEOFENCE CHECK
        distance = calculate_distance_meters(
            latitude, longitude,
            event.latitude, event.longitude
        )

        if distance > event.radius_meter:
            return JsonResponse({
                'error': f'Attendance rejected. Outside allowed area ({round(distance, 2)}m)'
            }, status=400)

        # ✅ SAVE ATTENDANCE (ANTI DUPLICATE)
        attendance, created = Attendance.objects.get_or_create(
            employee_id=employee_id,
            event=event,
            defaults={
                'full_name': employee.full_name,
                'phone_number': phone,
                'email': employee.email,
                'department': employee.department,
                'ipv4_address': ipv4_address,
                'ipv6_address': ipv6_address,
                'latitude': latitude,
                'longitude': longitude
            }
        )

        if not created:
            return JsonResponse({'error': 'Attendance already recorded'}, status=400)

        return JsonResponse({
            'message': 'Attendance successful',
            'distance_meter': round(distance, 2),
            'event': event.name
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def update_staff_attendance(request, id):
    permission_check = require_manage_api(request)
    if permission_check:
        return permission_check

    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        data = json.loads(request.body)
        att = Attendance.objects.get(id=id)

        att.full_name = data.get('full_name', att.full_name)
        att.employee_id = data.get('employee_id', att.employee_id)
        att.phone_number = data.get('phone_number', att.phone_number)
        att.email = data.get('email', att.email)
        att.department = data.get('department', att.department)
        att.save()

        return JsonResponse({'message': 'Staff attendance updated successfully'})
    except Attendance.DoesNotExist:
        return JsonResponse({'error': 'Staff attendance not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def delete_staff_attendance(request, id):
    permission_check = require_manage_api(request)
    if permission_check:
        return permission_check

    try:
        att = Attendance.objects.get(id=id)
        att.delete()
        return JsonResponse({'message': 'Staff attendance deleted successfully'})
    except Attendance.DoesNotExist:
        return JsonResponse({'error': 'Staff attendance not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def update_visitor_attendance(request, id):
    permission_check = require_manage_api(request)
    if permission_check:
        return permission_check

    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        data = json.loads(request.body)
        att = VisitorAttendance.objects.select_related('visitor').get(id=id)
        visitor = att.visitor

        visitor.full_name = data.get('full_name', visitor.full_name)
        visitor.phone_number = data.get('phone_number', visitor.phone_number)
        visitor.email = data.get('email', visitor.email)
        visitor.organization = data.get('organization', visitor.organization)
        visitor.save()

        return JsonResponse({'message': 'Visitor attendance updated successfully'})
    except VisitorAttendance.DoesNotExist:
        return JsonResponse({'error': 'Visitor attendance not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def delete_visitor_attendance(request, id):
    permission_check = require_manage_api(request)
    if permission_check:
        return permission_check

    try:
        att = VisitorAttendance.objects.get(id=id)
        att.delete()
        return JsonResponse({'message': 'Visitor attendance deleted successfully'})
    except VisitorAttendance.DoesNotExist:
        return JsonResponse({'error': 'Visitor attendance not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def update_passport_attendance(request, id):
    permission_check = require_manage_api(request)
    if permission_check:
        return permission_check

    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)

    try:
        data = json.loads(request.body)

        att = PassportAttendance.objects.select_related("passport_visitor").get(id=id)
        visitor = att.passport_visitor

        passport_type = (data.get("type") or "P").strip()
        country_code = (data.get("country_code") or "").strip().upper()
        passport_number = fix_common_ocr_errors(
            data.get("passport_number", ""),
            mode="passport"
        )
        nationality = (data.get("nationality") or "").strip()

        first_name = (data.get("first_name") or "").strip()
        last_name = (data.get("last_name") or "").strip()

        if not first_name and not last_name:
            first_name = (data.get("given_name") or "").strip()
            last_name = (data.get("surname") or "").strip()

        resolved_names = resolve_passport_name_parts({
            "first_name": first_name,
            "last_name": last_name,
            "country_code": country_code,
            "full_name": data.get("full_name", ""),
        })

        first_name = resolved_names["first_name"]
        last_name = resolved_names["last_name"]
        full_name = resolved_names["full_name"] or visitor.full_name

        date_of_birth = (data.get("date_of_birth") or "").strip()
        sex = (data.get("sex") or data.get("gender") or "").strip()
        date_of_issue = (data.get("date_of_issue") or "").strip()
        date_of_expiry = (data.get("date_of_expiry") or data.get("expiry_date") or "").strip()
        raw_text = (data.get("raw_text") or data.get("ocr_raw_text") or "").strip()
        status = (data.get("status") or visitor.status or "pending verification").strip()

        if not passport_number:
            return JsonResponse({"error": "Passport number cannot be empty"}, status=400)

        valid, passport_error = validate_passport_number_by_country(
            passport_number,
            country_code or nationality
        )
        if not valid:
            return JsonResponse({"error": passport_error}, status=400)

        additional_fields_text = normalize_additional_fields_text(
            data.get("additional_fields_text", "")
        )
        additional_fields_list = parse_additional_fields_text_to_list(additional_fields_text)

        existing_extra = dict(visitor.extra_data or {})

        visitor.full_name = full_name or visitor.full_name
        visitor.passport_number = passport_number
        visitor.country = nationality or country_code_to_name(country_code) or visitor.country
        visitor.date_of_birth = date_of_birth or ""
        visitor.expiry_date = date_of_expiry or ""
        visitor.gender = sex or ""
        visitor.ocr_raw_text = raw_text
        visitor.status = status

        visitor.extra_data = {
            **existing_extra,
            "type": passport_type,
            "country_code": country_code,
            "nationality": nationality,
            "first_name": first_name,
            "last_name": last_name,
            "date_of_issue": date_of_issue,
            "additional_fields_text": additional_fields_text,
            "additional_fields": additional_fields_list,
        }

        visitor.save()

        return JsonResponse({
            "message": "Non-Malaysian visitor updated successfully"
        })

    except PassportAttendance.DoesNotExist:
        return JsonResponse({"error": "Passport attendance not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def delete_passport_attendance(request, id):
    permission_check = require_manage_api(request)
    if permission_check:
        return permission_check

    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        att = PassportAttendance.objects.select_related('passport_visitor').get(id=id)
        visitor = att.passport_visitor

        att.delete()

        if not PassportAttendance.objects.filter(passport_visitor=visitor).exists():
            visitor.delete()

        return JsonResponse({'message': 'Passport attendance deleted successfully'})

    except PassportAttendance.DoesNotExist:
        return JsonResponse({'error': 'Passport attendance not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def event_detail(request, id):
    login_check = require_login_page(request)
    if login_check:
        return login_check

    event = get_object_or_404(Event, id=id)

    # =========================
    # STAFF ATTENDANCE
    # =========================
    staff_search = request.GET.get('staff_search', '').strip()
    staff_department = request.GET.get('staff_department', '').strip()
    staff_sort = request.GET.get('staff_sort', 'date')
    staff_page_num = request.GET.get('staff_page', 1)

    staff_qs = Attendance.objects.filter(event=event)

    if staff_search:
        staff_qs = staff_qs.filter(full_name__icontains=staff_search)

    if staff_department:
        staff_qs = staff_qs.filter(department__iexact=staff_department)

    if staff_sort == 'name':
        staff_qs = staff_qs.order_by('full_name')
    elif staff_sort == 'time':
        staff_qs = staff_qs.order_by('time')
    else:
        staff_qs = staff_qs.order_by('date', 'time')

    staff_departments = (
        Attendance.objects.filter(event=event)
        .exclude(department__isnull=True)
        .exclude(department__exact='')
        .values_list('department', flat=True)
        .distinct()
        .order_by('department')
    )

    staff_paginator = Paginator(staff_qs, 5)
    staff_page_obj = staff_paginator.get_page(staff_page_num)

    # =========================
    # VISITOR ATTENDANCE (MALAYSIAN)
    # =========================
    visitor_search = request.GET.get('visitor_search', '').strip()
    visitor_organization = request.GET.get('visitor_organization', '').strip()
    visitor_sort = request.GET.get('visitor_sort', 'date')
    visitor_page_num = request.GET.get('visitor_page', 1)

    visitor_qs = VisitorAttendance.objects.filter(event=event).select_related('visitor')

    if visitor_search:
        visitor_qs = visitor_qs.filter(visitor__full_name__icontains=visitor_search)

    if visitor_organization:
        visitor_qs = visitor_qs.filter(visitor__organization__iexact=visitor_organization)

    if visitor_sort == 'name':
        visitor_qs = visitor_qs.order_by('visitor__full_name')
    elif visitor_sort == 'time':
        visitor_qs = visitor_qs.order_by('time')
    else:
        visitor_qs = visitor_qs.order_by('date', 'time')

    visitor_organizations = (
        VisitorAttendance.objects.filter(event=event)
        .exclude(visitor__organization__isnull=True)
        .exclude(visitor__organization__exact='')
        .values_list('visitor__organization', flat=True)
        .distinct()
        .order_by('visitor__organization')
    )

    visitor_paginator = Paginator(visitor_qs, 5)
    visitor_page_obj = visitor_paginator.get_page(visitor_page_num)

    # =========================
    # PASSPORT ATTENDANCE (NON-MALAYSIAN)
    # =========================
    passport_search = request.GET.get('passport_search', '').strip()
    passport_country = request.GET.get('passport_country', '').strip()
    passport_sort = request.GET.get('passport_sort', 'date')
    passport_page_num = request.GET.get('passport_page', 1)

    passport_qs = PassportAttendance.objects.filter(event=event).select_related('passport_visitor')

    if passport_search:
        passport_qs = passport_qs.filter(passport_visitor__full_name__icontains=passport_search)

    if passport_country:
        passport_qs = passport_qs.filter(passport_visitor__country__iexact=passport_country)

    if passport_sort == 'name':
        passport_qs = passport_qs.order_by('passport_visitor__full_name')
    elif passport_sort == 'time':
        passport_qs = passport_qs.order_by('time')
    else:
        passport_qs = passport_qs.order_by('date', 'time')

    passport_countries = (
        PassportAttendance.objects.filter(event=event)
        .exclude(passport_visitor__country__isnull=True)
        .exclude(passport_visitor__country__exact='')
        .values_list('passport_visitor__country', flat=True)
        .distinct()
        .order_by('passport_visitor__country')
    )

    passport_paginator = Paginator(passport_qs, 5)
    passport_page_obj = passport_paginator.get_page(passport_page_num)

    for att in passport_page_obj.object_list:
        extra_data = att.passport_visitor.extra_data or {}
        additional_fields = extra_data.get("additional_fields", [])
        if not isinstance(additional_fields, list):
            additional_fields = []

        cleaned_additional_fields = []
        for field in additional_fields:
            if not isinstance(field, dict):
                continue
            label = str(field.get("label", "")).strip()
            value = str(field.get("value", "")).strip()
            if label and value:
                cleaned_additional_fields.append({
                    "label": label,
                    "value": value,
                })

        att.type_value = extra_data.get("type", "P")
        att.country_code_value = extra_data.get("country_code", "")
        att.nationality_value = extra_data.get(
            "nationality",
            att.passport_visitor.country or ""
        )
        att.first_name_value = extra_data.get("first_name", "")
        att.last_name_value = extra_data.get("last_name", "")
        att.date_of_issue_value = extra_data.get("date_of_issue", "")
        att.additional_fields_text = extra_data.get("additional_fields_text", "")
        att.additional_fields_json = json.dumps(cleaned_additional_fields, ensure_ascii=False)
        att.original_image_url = (
            att.passport_visitor.image.url
            if att.passport_visitor.image
            else ""
        )


    total_attendance = (
        staff_qs.count()
        + visitor_qs.count()
        + passport_qs.count()
    )

    context = {
        'event': event,
        'total_attendance': total_attendance,

        'employee_attendances': staff_page_obj.object_list,
        'staff_page_obj': staff_page_obj,
        'staff_departments': staff_departments,
        'staff_search': staff_search,
        'staff_department': staff_department,
        'staff_sort': staff_sort,

        'visitor_attendances': visitor_page_obj.object_list,
        'visitor_page_obj': visitor_page_obj,
        'visitor_organizations': visitor_organizations,
        'visitor_search': visitor_search,
        'visitor_organization': visitor_organization,
        'visitor_sort': visitor_sort,

        'passport_attendances': passport_page_obj.object_list,
        'passport_page_obj': passport_page_obj,
        'passport_countries': passport_countries,
        'passport_search': passport_search,
        'passport_country': passport_country,
        'passport_sort': passport_sort,
    }
    context.update(role_context(request))

    return render(request, 'event_detail.html', context)

def format_csv_date(value):
    if not value:
        return ""

    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")

    text = str(value).strip()
    if not text:
        return ""

    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", text)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"

    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", text)
    if m:
        return text

    m = re.match(r"^(\d{2})-(\d{2})-(\d{4})$", text)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"

    return text

def export_attendance_csv(request, id):
    manage_check = require_manage_page(request)
    if manage_check:
        return manage_check

    event = Event.objects.get(id=id)

    employee_attendances = Attendance.objects.filter(event=event)
    visitor_attendances = VisitorAttendance.objects.filter(event=event).select_related('visitor')
    passport_attendances = PassportAttendance.objects.filter(event=event).select_related('passport_visitor')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{event.name}_attendance.csv"'

    writer = csv.writer(response)

    # =========================
    # EMPLOYEE ATTENDANCE
    # =========================
    writer.writerow(['EMPLOYEE ATTENDANCE'])
    writer.writerow([
        'Name',
        'Employee ID',
        'Phone',
        'Email',
        'Department',
        'IPv4',
        'IPv6',
        'Date',
        'Time',
        'Latitude',
        'Longitude',
    ])

    for att in employee_attendances:
        writer.writerow([
            att.full_name,
            att.employee_id,
            f"'{att.phone_number}" if att.phone_number else '',
            att.email,
            att.department,
            getattr(att, 'ipv4_address', ''),
            getattr(att, 'ipv6_address', ''),
            att.date.strftime("%d/%m/%Y") if att.date else '',
            att.time.strftime("%H:%M:%S") if att.time else '',
            att.latitude,
            att.longitude,
        ])

    writer.writerow([])

    # =========================
    # VISITOR ATTENDANCE (MALAYSIAN)
    # =========================
    writer.writerow(['VISITOR ATTENDANCE (MALAYSIAN)'])
    writer.writerow([
        'Name',
        'Phone',
        'Email',
        'Organization',
        'Date',
        'Time',
        'Latitude',
        'Longitude',
    ])

    for att in visitor_attendances:
        writer.writerow([
            att.visitor.full_name,
            f"'{att.visitor.phone_number}" if att.visitor.phone_number else '',
            att.visitor.email,
            att.visitor.organization,
            att.date.strftime("%d/%m/%Y") if att.date else '',
            att.time.strftime("%H:%M:%S") if att.time else '',
            att.latitude,
            att.longitude,
        ])

    writer.writerow([])

    # =========================
    # VISITOR ATTENDANCE (NON-MALAYSIAN / PASSPORT)
    # =========================
    writer.writerow(['VISITOR ATTENDANCE (NON-MALAYSIAN / PASSPORT)'])
    writer.writerow([
        'Full Name',
        'First Name',
        'Last Name',
        'Passport Type',
        'Country Code',
        'Passport Number',
        'Nationality',
        'Country',
        'Date of Birth',
        'Sex',
        'Date of Issue',
        'Date of Expiry',
        'Status',
        'OCR Raw Text',
        'Additional Passport Fields',
        'Date',
        'Time',
        'Latitude',
        'Longitude',
    ])

    for att in passport_attendances:
        visitor = att.passport_visitor
        extra_data = visitor.extra_data or {}

        additional_source = extra_data.get('additional_fields')
        if not additional_source:
            additional_source = extra_data.get('additional_fields_text', '')

        cleaned_additional_fields = parse_additional_fields_text_to_list(additional_source)

        additional_fields_text = '; '.join(
            [f"{item.get('label', '')}: {item.get('value', '')}" for item in cleaned_additional_fields]
        )

        writer.writerow([
            visitor.full_name,
            extra_data.get('first_name', ''),
            extra_data.get('last_name', ''),
            extra_data.get('type', 'P'),
            extra_data.get('country_code', ''),
            visitor.passport_number,
            extra_data.get('nationality', visitor.country or ''),
            visitor.country,
            format_csv_date(visitor.date_of_birth),
            visitor.gender,
            format_csv_date(extra_data.get('date_of_issue', '')),
            format_csv_date(visitor.expiry_date),
            visitor.status,
            (visitor.ocr_raw_text or '').replace('\n', ' ').replace('\r', ' '),
            additional_fields_text,
            att.date.strftime("%d/%m/%Y") if att.date else '',
            att.time.strftime("%H:%M:%S") if att.time else '',
            att.latitude,
            att.longitude,
        ])

    return response


def export_event_summary_csv(request, id):
    manage_check = require_manage_page(request)
    if manage_check:
        return manage_check

    event = get_object_or_404(Event, id=id)

    staff_group = (
        Attendance.objects.filter(event=event)
        .exclude(department__isnull=True)
        .exclude(department__exact='')
        .values('department')
        .annotate(total=Count('id'))
        .order_by('-total', 'department')
    )

    visitor_group = (
        VisitorAttendance.objects.filter(event=event)
        .select_related('visitor')
        .exclude(visitor__organization__isnull=True)
        .exclude(visitor__organization__exact='')
        .values('visitor__organization')
        .annotate(total=Count('id'))
        .order_by('-total', 'visitor__organization')
    )

    passport_group = (
        PassportAttendance.objects.filter(event=event)
        .select_related('passport_visitor')
        .exclude(passport_visitor__country__isnull=True)
        .exclude(passport_visitor__country__exact='')
        .values('passport_visitor__country')
        .annotate(total=Count('id'))
        .order_by('-total', 'passport_visitor__country')
    )

    total_staff = Attendance.objects.filter(event=event).count()
    total_visitors = VisitorAttendance.objects.filter(event=event).count()
    total_passports = PassportAttendance.objects.filter(event=event).count()
    total_all = total_staff + total_visitors + total_passports

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{event.name}_summary.csv"'

    writer = csv.writer(response)

    writer.writerow(['EVENT SUMMARY'])
    writer.writerow(['Event Name', event.name])
    writer.writerow(['Start Date', format_csv_date(event.start_date)])
    writer.writerow(['End Date', format_csv_date(event.end_date)])
    writer.writerow(['Start Time', event.start_time.strftime("%H:%M") if event.start_time else '-'])
    writer.writerow(['End Time', event.end_time.strftime("%H:%M") if event.end_time else '-'])
    writer.writerow(['Location', event.location or '-'])
    writer.writerow(['Description', event.description if event.description else '-'])
    writer.writerow(['Latitude', event.latitude if event.latitude is not None else '-'])
    writer.writerow(['Longitude', event.longitude if event.longitude is not None else '-'])
    writer.writerow(['Radius (meter)', event.radius_meter if event.radius_meter is not None else '-'])
    writer.writerow([])

    writer.writerow(['ATTENDANCE TOTALS'])
    writer.writerow(['Category', 'Total'])
    writer.writerow(['Staff / Employee', total_staff])
    writer.writerow(['Visitor (Malaysian)', total_visitors])
    writer.writerow(['Visitor (Non-Malaysian / Passport)', total_passports])
    writer.writerow(['Overall Total', total_all])
    writer.writerow([])

    writer.writerow(['STAFF / EMPLOYEE BY DEPARTMENT'])
    writer.writerow(['Department', 'Total'])
    if staff_group:
        for item in staff_group:
            writer.writerow([item['department'], item['total']])
    else:
        writer.writerow(['-', 0])

    writer.writerow([])

    writer.writerow(['VISITOR (MALAYSIAN) BY ORGANIZATION'])
    writer.writerow(['Organization', 'Total'])
    if visitor_group:
        for item in visitor_group:
            writer.writerow([item['visitor__organization'], item['total']])
    else:
        writer.writerow(['-', 0])

    writer.writerow([])

    writer.writerow(['VISITOR (NON-MALAYSIAN / PASSPORT) BY COUNTRY'])
    writer.writerow(['Country', 'Total'])
    if passport_group:
        for item in passport_group:
            writer.writerow([item['passport_visitor__country'], item['total']])
    else:
        writer.writerow(['-', 0])

    return response

def ensure_media_dirs():
    os.makedirs(settings.MEDIA_ROOT, exist_ok=True)
    os.makedirs(os.path.join(settings.MEDIA_ROOT, 'passport_images'), exist_ok=True)
    os.makedirs(os.path.join(settings.MEDIA_ROOT, 'passport_processed'), exist_ok=True)

def get_country_choices():
    return sorted(set(COUNTRY_CODE_MAP.values()))


def normalize_display_date(date_str):
    if not date_str:
        return ""
    parsed = parse_passport_date(date_str)
    if not parsed:
        return ""
    return parsed.strftime("%Y-%m-%d")


def parse_passport_date(date_str):
    if not date_str:
        return None

    cleaned = date_str.strip().upper().replace("/", "-").replace(".", "-")
    cleaned = re.sub(r'\s+', ' ', cleaned)

    for fmt in [
        "%d %b %Y",
        "%d %B %Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
        "%d-%m-%y",
        "%d %b %y",
        "%d-%b-%Y",
        "%d-%b-%y",
    ]:
        try:
            return datetime.datetime.strptime(cleaned, fmt)
        except ValueError:
            pass

    return None


def is_expiry_valid(expiry_date_str):
    expiry_dt = parse_passport_date(expiry_date_str)
    if not expiry_dt:
        return False
    return expiry_dt.date() >= datetime.datetime.today().date()


def normalize_mrz_name(name):
    if not name:
        return ""

    name = fix_common_ocr_errors(name, mode="name")
    name = name.replace("<", " ")
    name = re.sub(r'\s+', ' ', name).strip()

    # split joined Malay/Japanese/English style chunks more safely
    name = re.sub(r'\bBIN(?=[A-Z])', 'BIN ', name)
    name = re.sub(r'\bBINTI(?=[A-Z])', 'BINTI ', name)

    return name.title()

def clean_name_token(value):
    if not value:
        return ""

    value = fix_common_ocr_errors(value, mode="name")
    value = value.replace("<", " ")
    value = re.sub(r"\s+", " ", value).strip()

    if not value:
        return ""

    return value.title()


def split_malaysian_name(full_name):
    """
    Malaysian rule:
    first_name = before BIN / BINTI
    last_name  = from BIN / BINTI until the end
    """
    if not full_name:
        return "", ""

    normalized = re.sub(r"\s+", " ", full_name).strip()
    upper_name = normalized.upper()

    match = re.search(r"\b(BIN|BINTI)\b", upper_name)
    if not match:
        return normalized, ""

    split_index = match.start()

    first_name = normalized[:split_index].strip()
    last_name = normalized[split_index:].strip()

    return first_name, last_name


def split_passport_name_parts(country_code, surname_part, given_part):
    """
    Generic MRZ rule:
    - surname_part => family name side
    - given_part   => given names side

    UI mapping:
    - first_name = given names
    - last_name  = family name

    Malaysian special handling:
    - if full name contains BIN/BINTI, split there
    """
    surname_clean = clean_name_token(surname_part)
    given_clean = clean_name_token(given_part)

    combined_full_name = " ".join(
        part for part in [given_clean, surname_clean] if part
    ).strip()

    if (country_code or "").upper() == "MYS":
        first_name, last_name = split_malaysian_name(combined_full_name)
        if first_name or last_name:
            return {
                "first_name": first_name,
                "last_name": last_name,
                "full_name": " ".join(
                    part for part in [first_name, last_name] if part
                ).strip(),
            }

    first_name = given_clean
    last_name = surname_clean

    full_name = " ".join(
        part for part in [first_name, last_name] if part
    ).strip()

    return {
        "first_name": first_name,
        "last_name": last_name,
        "full_name": full_name,
    }


def resolve_passport_name_parts(extracted):
    """
    Resolve name for UI + DB with clear fallback priority.

    Priority:
    1. explicit first_name / last_name
    2. surname / given_name from OCR or MRZ
    3. full_name fallback
    """
    country_code = (
        extracted.get("country_code")
        or extracted.get("nationality_code")
        or ""
    ).upper()

    first_name = (extracted.get("first_name") or "").strip()
    last_name = (extracted.get("last_name") or "").strip()

    if first_name or last_name:
        return {
            "first_name": first_name,
            "last_name": last_name,
            "full_name": " ".join(
                part for part in [first_name, last_name] if part
            ).strip(),
        }

    surname = (extracted.get("surname") or "").strip()
    given_name = (extracted.get("given_name") or "").strip()

    if surname or given_name:
        return split_passport_name_parts(
            country_code=country_code,
            surname_part=surname,
            given_part=given_name,
        )

    full_name = clean_name_token(extracted.get("full_name", ""))
    if not full_name:
        return {
            "first_name": "",
            "last_name": "",
            "full_name": "",
        }

    if country_code == "MYS":
        first_name, last_name = split_malaysian_name(full_name)
        return {
            "first_name": first_name,
            "last_name": last_name,
            "full_name": " ".join(
                part for part in [first_name, last_name] if part
            ).strip(),
        }

    parts = full_name.split()
    if len(parts) == 1:
        return {
            "first_name": parts[0],
            "last_name": "",
            "full_name": parts[0],
        }

    return {
        "first_name": " ".join(parts[:-1]).strip(),
        "last_name": parts[-1].strip(),
        "full_name": full_name,
    }


def get_safe_raw_text(extracted):
    return (
        extracted.get("raw_text")
        or extracted.get("ocr_raw_text")
        or extracted.get("mrz_text")
        or ""
    ).strip()

def repair_mrz_line2(line2):
    if not line2:
        return ""

    line2 = fix_common_ocr_errors(line2, mode="mrz")

    # common nationality OCR repair
    line2 = line2.replace("5PN", "JPN")
    line2 = line2.replace("7PN", "JPN")
    line2 = line2.replace("25PN", "JPN")
    line2 = line2.replace("MYS", "MYS")

    # try repair first 9 chars as passport no
    first9 = line2[:9]
    rest = line2[9:]

    first9 = re.sub(r'[^A-Z0-9<]', '', first9)
    first9 = first9.replace("O", "0").replace("Q", "0")
    first9 = first9.replace("I", "1").replace("L", "1")

    return first9 + rest

def parse_two_line_passport_mrz(line1, line2, rescue_mode=False):
    try:
        line1 = fix_common_ocr_errors(line1, mode="mrz")
        line2 = repair_mrz_line2(line2)

        line1 = re.sub(r"[^A-Z0-9<]", "", line1)
        line2 = re.sub(r"[^A-Z0-9<]", "", line2)

        line1 = line1[:44].ljust(44, "<")
        line2 = line2[:44].ljust(44, "<")

        if not line1.startswith("P<"):
            return None

        issuing_country = line1[2:5]
        names_part = line1[5:]

        surname_part = ""
        given_part = ""

        if "<<" in names_part:
            surname_part, given_part = names_part.split("<<", 1)
        else:
            repaired_names = names_part.replace("<K<", "<<")
            if "<<" not in repaired_names and "<" in repaired_names:
                repaired_names = repaired_names.replace("<", "<<", 1)

            if "<<" in repaired_names:
                surname_part, given_part = repaired_names.split("<<", 1)
            else:
                surname_part = repaired_names
                given_part = ""

        name_parts = split_passport_name_parts(
            country_code=issuing_country,
            surname_part=surname_part,
            given_part=given_part,
        )

        passport_raw = line2[0:9]
        passport_check = line2[9]
        nationality = line2[10:13]
        dob_raw = line2[13:19]
        dob_check = line2[19]
        gender_char = line2[20]
        expiry_raw = line2[21:27]
        expiry_check = line2[27]

        passport_number = fix_common_ocr_errors(
            passport_raw.replace("<", ""),
            mode="passport"
        )

        passport_valid = mrz_check_digit(passport_raw) == passport_check
        dob_valid = mrz_check_digit(dob_raw) == dob_check
        expiry_valid = mrz_check_digit(expiry_raw) == expiry_check

        def yyMMdd_to_ddmmyyyy(raw):
            if not raw.isdigit() or len(raw) != 6:
                return ""

            yy = int(raw[:2])
            mm = raw[2:4]
            dd = raw[4:6]

            current_year_2 = int(datetime.datetime.now().strftime("%y"))
            century = 2000 if yy <= current_year_2 else 1900

            return f"{dd}/{mm}/{century + yy}"

        date_of_birth = yyMMdd_to_ddmmyyyy(dob_raw) if (dob_valid or rescue_mode) else ""
        date_of_expiry = yyMMdd_to_ddmmyyyy(expiry_raw) if (expiry_valid or rescue_mode) else ""

        sex = {"M": "Male", "F": "Female"}.get(gender_char, gender_char)

        return {
            "type": "P",
            "country_code": issuing_country,
            "passport_number": passport_number,
            "nationality": country_code_to_name(nationality),
            "nationality_code": nationality,
            "first_name": name_parts["first_name"],
            "last_name": name_parts["last_name"],
            "full_name": name_parts["full_name"],
            "date_of_birth": date_of_birth,
            "sex": sex,
            "date_of_expiry": date_of_expiry,
            "raw_text": f"{line1}\n{line2}",
            "status": "auto-extracted" if passport_valid else "pending verification",
            "confidence_score": 95 if passport_valid else 70,
        }

    except Exception:
        return None

def correct_image_rotation(image):
    if image is None:
        return image

    # keep simple and safe first
    h, w = image.shape[:2]
    if h > w * 1.35:
        image = cv2.rotate(image, cv2.ROTATE_90_CLOCKWISE)

    return image

def fix_upside_down(image):
    try:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # split top & bottom
        h = gray.shape[0]
        top = gray[:h//2, :]
        bottom = gray[h//2:, :]

        # calculate brightness
        top_mean = np.mean(top)
        bottom_mean = np.mean(bottom)

        # MRZ usually darker (bottom)
        if top_mean < bottom_mean:
            image = cv2.rotate(image, cv2.ROTATE_180)

        return image

    except Exception as e:
        print("Upside-down detection error:", e)
        return image
    
def preprocess_passport_image(image_path, output_path):
    image = cv2.imread(image_path)

    # 1. straighten angle
    image = auto_rotate_passport(image)

    # 2. fix upside down
    image = fix_upside_down(image)
    if image is None:
        raise Exception("Failed to read uploaded passport image")

    image = correct_image_rotation(image)
    h, w = image.shape[:2]

    # light crop only
    cropped = image[int(h * 0.02):int(h * 0.98), int(w * 0.02):int(w * 0.98)]
    resized = cv2.resize(cropped, None, fx=1.8, fy=1.8, interpolation=cv2.INTER_CUBIC)

    gray = cv2.cvtColor(resized, cv2.COLOR_BGR2GRAY)
    denoise = cv2.bilateralFilter(gray, 7, 50, 50)
    enhanced = cv2.convertScaleAbs(denoise, alpha=1.25, beta=10)

    cv2.imwrite(output_path, enhanced)

    blur_value = cv2.Laplacian(gray, cv2.CV_64F).var()
    image_quality_note = None
    if blur_value < 60:
        image_quality_note = "Image may be too blurry. Please upload a clearer passport image."

    return enhanced, image_quality_note


def extract_passport_data_from_text(text):
    surname = ""
    given_name = ""
    date_of_birth = ""
    date_of_issue = ""
    date_of_expiry = ""
    nationality = ""
    sex = ""
    passport_number = "" 
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    upper_lines = [line.upper() for line in lines]
    full_upper = "\n".join(upper_lines) 

    result = {
        'type': 'P',
        'full_name': '',
        'surname': '',
        'given_name': '',
        'passport_number': '',
        'country': '',
        'country_code': '',
        'nationality': '',
        'date_of_birth': '',
        'date_of_issue': '',
        'expiry_date': '',
        'gender': '',
        'sex': '',
        'registered_domicile': '',
        'issuing_authority': '',
    }

    # passport number visual fallback
    passport_match = re.search(r'\b[A-Z]{1,2}\d{6,8}\b', full_upper)
    if passport_match:
        result['passport_number'] = fix_common_ocr_errors(passport_match.group(0), mode="passport")

    # country code
    for code in COUNTRY_CODE_MAP.keys():
        if f" {code} " in f" {full_upper} ":
            result['country_code'] = code
            result['country'] = country_code_to_name(code)
            result['nationality'] = country_code_to_name(code)
            break

    # safer label parsing
    for i, line in enumerate(upper_lines):
        next_line = upper_lines[i + 1] if i + 1 < len(upper_lines) else ""

        if ('DATE OF BIRTH' in line or 'DATE OF BIRT' in line or 'BIRTH' in line) and next_line:
            dob_match = re.search(r'\b\d{1,2}\s+[A-Z]{3}\s+\d{4}\b', next_line)
            if dob_match:
                result['date_of_birth'] = dob_match.group(0)

        if ('DATE OF ISSUE' in line or 'ISSUE' in line) and next_line:
            issue_match = re.search(r'\b\d{1,2}\s+[A-Z]{3}\s+\d{4}\b', next_line)
            if issue_match:
                result['date_of_issue'] = issue_match.group(0)

        if ('DATE OF EXPIRY' in line or 'EXPIRY' in line or 'EXPIRE' in line) and next_line:
            exp_match = re.search(r'\b\d{1,2}\s+[A-Z]{3}\s+\d{4}\b', next_line)
            if exp_match:
                result['expiry_date'] = exp_match.group(0)

        if ('SURNAME' in line or 'S/SURNAME' in line) and next_line:
            candidate = re.sub(r'[^A-Z\s]', ' ', next_line).strip()
            candidate = re.sub(r'\s+', ' ', candidate)
            if candidate and len(candidate.split()) <= 3:
                result['surname'] = normalize_mrz_name(candidate)

        if ('GIVEN NAME' in line or 'G/GIVEN NAME' in line or 'GIVENNAME' in line) and next_line:
            candidate = re.sub(r'[^A-Z\s]', ' ', next_line).strip()
            candidate = re.sub(r'\s+', ' ', candidate)
            if candidate and len(candidate.split()) <= 4:
                result['given_name'] = normalize_mrz_name(candidate)

        if 'REGISTERED DOMICILE' in line and next_line:
            candidate = re.sub(r'[^A-Z\s]', ' ', next_line).strip()
            candidate = re.sub(r'\s+', ' ', candidate)
            if candidate:
                result['registered_domicile'] = normalize_mrz_name(candidate)

        if 'AUTHORITY' in line and next_line:
            candidate = re.sub(r'[^A-Z\s]', ' ', next_line).strip()
            candidate = re.sub(r'\s+', ' ', candidate)
            if candidate:
                result['issuing_authority'] = normalize_mrz_name(candidate)

    # dates visible area
    dates = re.findall(r'\b\d{1,2}\s+[A-Z]{3}\s+\d{4}\b', full_upper)
    unique_dates = []
    for d in dates:
        if d not in unique_dates:
            unique_dates.append(d)

    if len(unique_dates) >= 1:
        result['date_of_birth'] = unique_dates[0]
    if len(unique_dates) >= 2:
        result['date_of_issue'] = unique_dates[1]
    if len(unique_dates) >= 3:
        result['expiry_date'] = unique_dates[2]

    surname = clean_person_name(surname)
    given_name = clean_person_name(given_name)

    date_of_birth = normalize_display_date(date_of_birth)
    date_of_issue = normalize_display_date(date_of_issue)
    date_of_expiry = normalize_display_date(date_of_expiry)

    # only fallback from unique_dates if field still empty
    if not date_of_birth and len(unique_dates) >= 1:
        date_of_birth = normalize_display_date(unique_dates[0])

    if not date_of_issue and len(unique_dates) >= 2:
        date_of_issue = normalize_display_date(unique_dates[1])

    if not date_of_expiry and len(unique_dates) >= 3:
        date_of_expiry = normalize_display_date(unique_dates[2])

    # sex visual fallback - stricter
    sex_match = re.search(r'\bSEX\b[\s:]*([MF])\b', full_upper)
    if sex_match:
        if sex_match.group(1) == 'F':
            result['gender'] = 'Female'
            result['sex'] = 'Female'
        elif sex_match.group(1) == 'M':
            result['gender'] = 'Male'
            result['sex'] = 'Male'

    if not result['issuing_authority'] and 'MINISTRY OF FOREIGN AFFAIRS' in full_upper:
        result['issuing_authority'] = 'Ministry of Foreign Affairs'

    result['full_name'] = f"{result['surname']} {result['given_name']}".strip()

    return result

def merge_passport_results(mrz_result, visual_result):
    final_result = {}

    mrz_result = mrz_result or {}
    visual_result = visual_result or {}

    # MRZ fields = highest priority
    final_result["type"] = mrz_result.get("type") or visual_result.get("type") or "P"
    final_result["passport_number"] = mrz_result.get("passport_number") or visual_result.get("passport_number", "")
    final_result["surname"] = mrz_result.get("surname") or visual_result.get("surname", "")
    final_result["given_name"] = mrz_result.get("given_name") or visual_result.get("given_name", "")
    final_result["full_name"] = mrz_result.get("full_name") or visual_result.get("full_name", "")

    final_result["country"] = mrz_result.get("country") or visual_result.get("country", "")
    final_result["country_code"] = mrz_result.get("country_code") or visual_result.get("country_code", "")
    final_result["nationality"] = mrz_result.get("nationality") or visual_result.get("nationality", "")
    final_result["nationality_code"] = mrz_result.get("nationality_code") or visual_result.get("nationality_code", "")

    final_result["date_of_birth"] = mrz_result.get("date_of_birth") or visual_result.get("date_of_birth", "")
    final_result["expiry_date"] = mrz_result.get("expiry_date") or visual_result.get("expiry_date", "")

    final_result["sex"] = mrz_result.get("sex") or visual_result.get("sex", "")
    final_result["gender"] = mrz_result.get("gender") or visual_result.get("gender", "")

    # visual fallback only
    final_result["date_of_issue"] = visual_result.get("date_of_issue", "")
    final_result["registered_domicile"] = visual_result.get("registered_domicile", "")
    final_result["issuing_authority"] = visual_result.get("issuing_authority", "")

    final_result["mrz_valid_score"] = mrz_result.get("mrz_valid_score", 0)
    final_result["mrz_total_checks"] = mrz_result.get("mrz_total_checks", 3)
    final_result["raw_text"] = visual_result.get("raw_text", "")
    final_result["avg_confidence"] = max(
        mrz_result.get("avg_confidence", 0),
        visual_result.get("avg_confidence", 0)
    )

    return final_result

def score_extraction_result(result):
    if not result:
        return -1

    score = 0

    if result.get("passport_number"):
        score += 4
    if result.get("surname"):
        score += 3
    if result.get("given_name"):
        score += 3
    if result.get("nationality"):
        score += 2
    if result.get("date_of_birth"):
        score += 2
    if result.get("expiry_date"):
        score += 2
    if result.get("sex"):
        score += 1
    if result.get("date_of_issue"):
        score += 1
    if result.get("registered_domicile"):
        score += 1
    if result.get("issuing_authority"):
        score += 1

    score += result.get("mrz_valid_score", 0) * 6
    score += int(result.get("avg_confidence", 0) * 5)

    # penalize obvious junk
    for field in ["passport_number", "surname", "given_name", "nationality"]:
        val = str(result.get(field, "")).strip().upper()
        if val and re.search(r'\d{3,}', val):
            score -= 3

    return score

def run_paddleocr_retry_variants(input_img, request_id=None):
    if not request_id:
        request_id = uuid.uuid4().hex

    results = []

    full_lines = paddleocr_lines_from_image(input_img, f"full_passport_{request_id}.jpg")
    full_text = "\n".join([x["text"] for x in full_lines])
    full_avg = sum([x["score"] for x in full_lines]) / len(full_lines) if full_lines else 0

    visual_result = extract_passport_data_from_text(full_text)
    visual_result["raw_text"] = full_text
    visual_result["avg_confidence"] = full_avg

    for filename, candidate_img in get_mrz_variants(input_img):
        safe_filename = f"{request_id}_{filename}"
        mrz_lines = paddleocr_lines_from_image(candidate_img, safe_filename)
        mrz_text = "\n".join([x["text"] for x in mrz_lines])
        mrz_avg = sum([x["score"] for x in mrz_lines]) / len(mrz_lines) if mrz_lines else 0

        mrz_data = extract_mrz_data(mrz_text)
        if mrz_data:
            merged = merge_passport_results(mrz_data, visual_result)
            merged["raw_text"] = full_text
            merged["avg_confidence"] = max(mrz_avg, full_avg)
            results.append(merged)

    rescue_full = parse_mrz_rescue(full_text)
    if rescue_full:
        merged_rescue = merge_passport_results(rescue_full, visual_result)
        merged_rescue["raw_text"] = full_text
        merged_rescue["avg_confidence"] = full_avg
        results.append(merged_rescue)

    visual_result["raw_text"] = full_text
    results.append(visual_result)

    if not results:
        return visual_result, score_extraction_result(visual_result)

    best_result = max(results, key=score_extraction_result)
    best_score = score_extraction_result(best_result)
    return best_result, best_score

def rotate_image_by_angle(image, angle):
    if image is None:
        return None

    angle = int(angle) % 360

    if angle == 0:
        return image.copy()
    if angle == 90:
        return cv2.rotate(image, cv2.ROTATE_90_CLOCKWISE)
    if angle == 180:
        return cv2.rotate(image, cv2.ROTATE_180)
    if angle == 270:
        return cv2.rotate(image, cv2.ROTATE_90_COUNTERCLOCKWISE)

    return image.copy()


def choose_best_orientation_by_ocr(image, request_id=None):
    if image is None:
        raise Exception("Image is empty for orientation detection")

    if not request_id:
        request_id = uuid.uuid4().hex

    orientation_candidates = [0, 90, 180, 270]
    best_angle = 0
    best_score = -999999
    best_result = None
    best_rotated_image = image.copy()

    for angle in orientation_candidates:
        rotated_img = rotate_image_by_angle(image, angle)

        try:
            result, score = run_paddleocr_retry_variants(
                rotated_img,
                request_id=f"{request_id}_rot{angle}"
            )
        except TypeError:
            # fallback kalau function lama tak ada request_id param
            result, score = run_paddleocr_retry_variants(rotated_img)

        # tambah sedikit bonus kalau OCR text nampak lebih "passport-like"
        raw_text = (result or {}).get("raw_text", "") or ""
        upper_text = raw_text.upper()

        orientation_bonus = 0
        if "PASSPORT" in upper_text:
            orientation_bonus += 2
        if "P<" in upper_text:
            orientation_bonus += 3
        if result and result.get("passport_number"):
            orientation_bonus += 3
        if result and result.get("surname"):
            orientation_bonus += 2
        if result and result.get("given_name"):
            orientation_bonus += 2
        if result and result.get("mrz_valid_score", 0) >= 1:
            orientation_bonus += 5

        total_score = score + orientation_bonus

        print(f"[ORIENTATION CHECK] angle={angle} score={score} bonus={orientation_bonus} total={total_score}")

        if total_score > best_score:
            best_score = total_score
            best_angle = angle
            best_result = result
            best_rotated_image = rotated_img

    return best_rotated_image, best_result, best_score, best_angle

def process_passport_ocr(image_path, processed_path, request_id=None):
    if not request_id:
        request_id = uuid.uuid4().hex

    image_quality_note = None

    original_img = cv2.imread(image_path)
    if original_img is None:
        raise Exception("Failed to read uploaded passport image")

    # Step 1: OCR orientation check sekali sahaja pada gambar asal
    oriented_original_img, original_result, original_score, detected_angle = choose_best_orientation_by_ocr(
        original_img,
        request_id=f"{request_id}_original"
    )

    print(f"[BEST ORIGINAL ORIENTATION] angle={detected_angle} score={original_score}")

    if oriented_original_img is None:
        raise Exception("Failed to orient passport image")

    # Simpan oriented RGB image dulu
    cv2.imwrite(processed_path, oriented_original_img)

    # Step 2: preprocess hanya untuk simpan processed file / optional support
    processed_img, image_quality_note = preprocess_passport_image(processed_path, processed_path)

    # Step 3: JANGAN run 4 orientation lagi pada processed image
    best_result = original_result
    best_score = original_score
    best_angle = detected_angle

    # Simpan processed image dalam folder media/passport_processed
    if processed_img is not None:
        cv2.imwrite(processed_path, processed_img)

    if not best_result:
        raise Exception("OCR could not detect any text from the passport image")

    status = 'pending verification'
    if best_result.get("mrz_valid_score", 0) >= 2:
        status = 'auto-extracted'

    best_result['status'] = status
    best_result['image_quality_note'] = image_quality_note
    best_result['confidence_score'] = best_score
    best_result['detected_rotation_angle'] = best_angle
    best_result['date_of_birth'] = normalize_display_date(best_result.get('date_of_birth'))
    best_result['date_of_issue'] = normalize_display_date(best_result.get('date_of_issue'))
    best_result['expiry_date'] = normalize_display_date(best_result.get('expiry_date'))

    return best_result

def paddleocr_lines_from_image(image, temp_name=None):
    lines = []

    try:
        ensure_media_dirs()

        if not temp_name:
            temp_name = f"temp_{uuid.uuid4().hex}.jpg"

        temp_path = os.path.join(settings.MEDIA_ROOT, "passport_processed", temp_name)

        if isinstance(image, np.ndarray):
            cv2.imwrite(temp_path, image)
            predict_input = temp_path
        else:
            predict_input = image

        # TRY 1: predict()
        try:
            with PADDLE_OCR_LOCK:
                result = PADDLE_OCR.predict(input=predict_input)

            for item in result:
                item_json = None

                if hasattr(item, "json"):
                    raw_json = item.json
                    if isinstance(raw_json, str):
                        try:
                            item_json = json.loads(raw_json)
                        except Exception:
                            item_json = None
                    elif isinstance(raw_json, dict):
                        item_json = raw_json
                elif isinstance(item, dict):
                    item_json = item

                if not item_json:
                    continue

                rec_texts = item_json.get("rec_texts", []) or []
                rec_scores = item_json.get("rec_scores", []) or []

                for i, txt in enumerate(rec_texts):
                    score = rec_scores[i] if i < len(rec_scores) else 0.0
                    if txt and str(txt).strip():
                        lines.append({
                            "text": str(txt).strip(),
                            "score": float(score) if score is not None else 0.0
                        })
        except Exception:
            pass

        # TRY 2: ocr() fallback
        if not lines:
            try:
                with PADDLE_OCR_LOCK:
                    ocr_result = PADDLE_OCR.ocr(predict_input, cls=False)

                if isinstance(ocr_result, list):
                    for block in ocr_result:
                        if not block:
                            continue

                        for line in block:
                            if isinstance(line, list) and len(line) >= 2:
                                rec_part = line[1]

                                if isinstance(rec_part, (list, tuple)) and len(rec_part) >= 2:
                                    text = rec_part[0]
                                    score = rec_part[1]

                                    if text and str(text).strip():
                                        lines.append({
                                            "text": str(text).strip(),
                                            "score": float(score) if score is not None else 0.0
                                        })
            except Exception:
                pass

        # TRY 3: TESSERACT fallback
        if not lines:
            try:
                if isinstance(image, np.ndarray):
                    text = pytesseract.image_to_string(image)
                else:
                    text = pytesseract.image_to_string(cv2.imread(predict_input))

                if text and text.strip():
                    for row in text.split("\n"):
                        row = row.strip()
                        if row:
                            lines.append({
                                "text": row,
                                "score": 0.5
                            })
            except Exception:
                pass

    except Exception:
        return []

    return lines

def auto_rotate_passport(image):
    try:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # detect edges
        edges = cv2.Canny(gray, 50, 150)

        # detect lines
        lines = cv2.HoughLines(edges, 1, np.pi / 180, 200)

        if lines is None:
            return image

        angles = []

        for rho, theta in lines[:, 0]:
            angle = (theta * 180 / np.pi) - 90
            angles.append(angle)

        if not angles:
            return image

        median_angle = np.median(angles)

        # rotate image
        (h, w) = image.shape[:2]
        center = (w // 2, h // 2)

        M = cv2.getRotationMatrix2D(center, median_angle, 1.0)
        rotated = cv2.warpAffine(image, M, (w, h),
                                 flags=cv2.INTER_CUBIC,
                                 borderMode=cv2.BORDER_REPLICATE)

        return rotated

    except Exception as e:
        print("Rotation error:", e)
        return image

def extract_mrz_data(text):
    text = fix_common_ocr_errors(text, mode="mrz")
    lines = [line.strip() for line in text.split('\n') if line.strip()]

    normalized = []
    for line in lines:
        clean_line = re.sub(r'[^A-Z0-9<]', '', line.upper())
        if len(clean_line) >= 25:
            normalized.append(clean_line)

    # try adjacent pairs
    for i in range(len(normalized) - 1):
        parsed = parse_two_line_passport_mrz(normalized[i], normalized[i + 1])
        if parsed:
            return parsed

    rescue = parse_mrz_rescue(text)
    if rescue:
        return rescue

    return None


def crop_mrz_region(image):
    h, w = image.shape[:2]
    return image[int(h * 0.72):h, 0:w]

def get_mrz_variants(image):
    mrz = crop_mrz_region(image)
    unique = uuid.uuid4().hex

    variants = []

    variants.append((f"mrz_raw_{unique}.jpg", mrz))

    gray = cv2.cvtColor(mrz, cv2.COLOR_BGR2GRAY) if len(mrz.shape) == 3 else mrz
    variants.append((f"mrz_gray_{unique}.jpg", gray))

    big = cv2.resize(gray, None, fx=3.0, fy=3.0, interpolation=cv2.INTER_CUBIC)
    variants.append((f"mrz_big_{unique}.jpg", big))

    _, thresh = cv2.threshold(big, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    variants.append((f"mrz_thresh_{unique}.jpg", thresh))

    inv = cv2.bitwise_not(thresh)
    variants.append((f"mrz_inv_{unique}.jpg", inv))

    return variants

def passport_attendance_page(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    return render(request, 'passport_attendance_form.html', {
        'event': event,
        'country_choices': get_country_choices(),
    })


@csrf_exempt
def upload_passport(request):
    try:
        if request.method != "POST":
            return JsonResponse({"error": "Invalid request method"}, status=400)

        if "image" not in request.FILES:
            return JsonResponse({"error": "Please choose a passport image first"}, status=400)

        ensure_media_dirs()

        image = request.FILES["image"]
        ext = os.path.splitext(image.name)[1].lower() or ".jpg"
        unique_id = uuid.uuid4().hex

        original_filename = f"passport_{unique_id}{ext}"
        processed_filename = f"processed_{unique_id}.jpg"

        original_path = os.path.join(settings.MEDIA_ROOT, "passport_images", original_filename)
        processed_path = os.path.join(settings.MEDIA_ROOT, "passport_processed", processed_filename)

        with open(original_path, "wb+") as file_obj:
            for chunk in image.chunks():
                file_obj.write(chunk)

        extracted = process_passport_ocr(original_path, processed_path)
        ui_result = build_universal_passport_fields(extracted)

        if not ui_result.get("first_name") and extracted.get("given_name"):
            ui_result["first_name"] = extracted.get("given_name", "")

        if not ui_result.get("last_name") and extracted.get("surname"):
            ui_result["last_name"] = extracted.get("surname", "")

        if not ui_result.get("sex") and extracted.get("gender"):
            ui_result["sex"] = extracted.get("gender", "")

        final_status = ui_result.get("status", "pending verification")
        if not ui_result.get("passport_number"):
            final_status = "pending verification"

        return JsonResponse({
            "message": "Passport scanned successfully",
            "type": ui_result.get("type", "P"),
            "country_code": ui_result.get("country_code", ""),
            "passport_number": ui_result.get("passport_number", ""),
            "nationality": ui_result.get("nationality", ""),
            "first_name": ui_result.get("first_name", ""),
            "last_name": ui_result.get("last_name", ""),
            "full_name": ui_result.get("full_name", ""),
            "date_of_birth": normalize_display_date(ui_result.get("date_of_birth", "")),
            "sex": ui_result.get("sex", ""),
            "date_of_issue": normalize_display_date(ui_result.get("date_of_issue", "")),
            "date_of_expiry": normalize_display_date(
                ui_result.get("date_of_expiry", "")
            ),
            "raw_text": ui_result.get("raw_text", ""),
            "status": final_status,
            "confidence_score": ui_result.get("confidence_score", 0),
            "image_quality_note": ui_result.get("image_quality_note", ""),
            "detected_rotation_angle": ui_result.get("detected_rotation_angle", 0),
            "original_image_name": original_filename,
            "processed_image_name": processed_filename,
            "original_image_url": f"{settings.MEDIA_URL}passport_images/{original_filename}",
            "processed_image_url": f"{settings.MEDIA_URL}passport_processed/{processed_filename}",
        })

    except Exception as e:
        return JsonResponse({
            "error": str(e),
            "status": "pending verification",
        }, status=500)



def normalize_additional_fields_text(value):
    if value is None:
        return ""

    text = str(value)

    # Handle literal escaped sequences yang mungkin tersimpan dalam DB
    text = text.replace("\\u000A", "\n").replace("\\u000a", "\n")
    text = text.replace("\\r\\n", "\n").replace("\\n", "\n")

    # Normalise carriage returns
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    lines = []
    for raw_line in text.split("\n"):
        line = raw_line.strip()
        if line:
            lines.append(line)

    return "\n".join(lines)


def parse_additional_fields_text_to_list(value):
    text = normalize_additional_fields_text(value)
    if not text:
        return []

    result = []

    for raw_line in text.split("\n"):
        line = raw_line.strip()
        if not line:
            continue

        if ":" in line:
            label, field_value = line.split(":", 1)
            label = label.strip()
            field_value = field_value.strip()

            if label and field_value:
                result.append({
                    "label": label,
                    "value": field_value,
                })
        else:
            result.append({
                "label": "Note",
                "value": line,
            })

    return result
@csrf_exempt
def submit_passport_attendance(request, event_id):
    try:
        if request.method != "POST":
            return JsonResponse({"error": "Invalid request"}, status=400)

        data = json.loads(request.body)

        passport_type = (data.get("type") or "P").strip()
        country_code = (data.get("country_code") or "").strip().upper()
        passport_number = fix_common_ocr_errors(
            data.get("passport_number", ""),
            mode="passport"
        )
        nationality = (data.get("nationality") or "").strip()

        first_name = (data.get("first_name") or "").strip()
        last_name = (data.get("last_name") or "").strip()

        if not first_name and not last_name:
            first_name = (data.get("given_name") or "").strip()
            last_name = (data.get("surname") or "").strip()

        resolved_names = resolve_passport_name_parts({
            "first_name": first_name,
            "last_name": last_name,
            "country_code": country_code,
            "full_name": data.get("full_name", ""),
        })

        first_name = resolved_names["first_name"]
        last_name = resolved_names["last_name"]
        full_name = resolved_names["full_name"] or passport_number

        date_of_birth = (data.get("date_of_birth") or "").strip()
        sex = (data.get("sex") or data.get("gender") or "").strip()
        date_of_issue = (data.get("date_of_issue") or "").strip()
        date_of_expiry = (data.get("date_of_expiry") or data.get("expiry_date") or "").strip()

        raw_text = data.get("raw_text") or ""
        status = (data.get("status") or "pending verification").strip()

        original_image_name = (data.get("original_image_name") or "").strip()
        processed_image_name = (data.get("processed_image_name") or "").strip()

        latitude = data.get("latitude")
        longitude = data.get("longitude")

        # VALIDATION
        if not passport_number:
            return JsonResponse({"error": "Passport number cannot be empty"}, status=400)

        valid, passport_error = validate_passport_number_by_country(
            passport_number,
            country_code or nationality,
        )
        if not valid:
            return JsonResponse({"error": passport_error}, status=400)

        if latitude in [None, ""] or longitude in [None, ""]:
            return JsonResponse({"error": "Please enable GPS/location first"}, status=400)

        event = get_object_or_404(Event, id=event_id)

        if event.latitude is None or event.longitude is None:
            return JsonResponse({"error": "Event location not configured by admin"}, status=400)

        distance = calculate_distance_meters(
            latitude,
            longitude,
            event.latitude,
            event.longitude,
        )

        if distance > event.radius_meter:
            return JsonResponse({
                "error": f"Attendance rejected. Outside allowed area ({round(distance, 2)}m)"
            }, status=400)

        # CREATE / UPDATE VISITOR
        visitor, created = PassportVisitor.objects.get_or_create(
            passport_number=passport_number,
            defaults={
                "full_name": full_name,
                "country": nationality or country_code_to_name(country_code),
                "date_of_birth": date_of_birth,
                "expiry_date": date_of_expiry,
                "gender": sex,
                "ocr_raw_text": raw_text,
                "status": status,
                "extra_data": {
                    "type": passport_type,
                    "country_code": country_code,
                    "nationality": nationality,
                    "first_name": first_name,
                    "last_name": last_name,
                    "date_of_issue": date_of_issue,
                },
            },
        )

        if not created:
            visitor.full_name = full_name or visitor.full_name
            visitor.country = nationality or country_code_to_name(country_code) or visitor.country
            visitor.date_of_birth = date_of_birth or visitor.date_of_birth
            visitor.expiry_date = date_of_expiry or visitor.expiry_date
            visitor.gender = sex or visitor.gender
            visitor.ocr_raw_text = raw_text or visitor.ocr_raw_text
            visitor.status = status or visitor.status

        # ✅ FIXED JSON STRUCTURE
        merged_extra = dict(visitor.extra_data or {})

        additional_fields_text = normalize_additional_fields_text(
            data.get("additional_fields_text", "")
        )
        submit_additional = parse_additional_fields_text_to_list(
            additional_fields_text
        )

        merged_extra.update({
            "type": passport_type,
            "country_code": country_code,
            "nationality": nationality,
            "first_name": first_name,
            "last_name": last_name,
            "date_of_issue": date_of_issue,
            "additional_fields_text": additional_fields_text,
            "additional_fields": submit_additional,
        })

        visitor.extra_data = merged_extra

        # IMAGE
        if original_image_name:
            original_path = os.path.join(settings.MEDIA_ROOT, "passport_images", original_image_name)
            if os.path.exists(original_path):
                visitor.image.name = f"passport_images/{original_image_name}"

        if processed_image_name:
            processed_path = os.path.join(settings.MEDIA_ROOT, "passport_processed", processed_image_name)
            if os.path.exists(processed_path):
                visitor.extracted_image.name = f"passport_processed/{processed_image_name}"

        visitor.save()

        # ATTENDANCE
        attendance, attendance_created = PassportAttendance.objects.get_or_create(
            passport_visitor=visitor,
            event=event,
            defaults={
                "latitude": latitude,
                "longitude": longitude,
            },
        )

        if not attendance_created:
            return JsonResponse({"error": "Attendance already recorded"}, status=400)

        return JsonResponse({
            "message": "Passport attendance recorded successfully",
            "distance_meter": round(distance, 2),
            "event": event.name,
            "status": visitor.status,
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def mrz_char_value(c):
    if c.isdigit():
        return int(c)
    if 'A' <= c <= 'Z':
        return ord(c) - 55
    if c == '<':
        return 0
    return 0


def mrz_check_digit(data):
    weights = [7, 3, 1]
    total = 0
    for i, char in enumerate(data):
        total += mrz_char_value(char) * weights[i % 3]
    return str(total % 10)

def build_universal_passport_fields(extracted):
    name_parts = resolve_passport_name_parts(extracted)

    raw_text = get_safe_raw_text(extracted)

    sex_value = (
        extracted.get("sex")
        or extracted.get("gender")
        or extracted.get("extra_data", {}).get("sex")
        or extracted.get("extra_data", {}).get("gender")
        or ""
    )

    result = {
        "type": extracted.get("type", "P"),
        "country_code": extracted.get("country_code", extracted.get("nationality_code", "")),
        "passport_number": extracted.get("passport_number", ""),
        "first_name": name_parts["first_name"],
        "last_name": name_parts["last_name"],
        "full_name": name_parts["full_name"],
        "nationality": extracted.get("nationality", extracted.get("country", "")),
        "date_of_birth": extracted.get("date_of_birth", ""),
        "sex": sex_value,
        "date_of_issue": extracted.get("date_of_issue", ""),
        "date_of_expiry": extracted.get("date_of_expiry", extracted.get("expiry_date", "")),
        "raw_text": raw_text,
        "status": extracted.get("status", "pending verification"),
        "confidence_score": extracted.get("confidence_score", 0),
        "image_quality_note": extracted.get("image_quality_note", ""),
        "detected_rotation_angle": extracted.get("detected_rotation_angle", 0),
    }

    return result

def merge_passport_results(primary_result, fallback_result):
    if not primary_result:
        return fallback_result or {}

    final_result = dict(primary_result)

    if not fallback_result:
        return final_result

    fields_from_fallback = [
        # new naming structure
        "first_name",
        "last_name",
        "full_name",

        # old naming structure, keep for compatibility
        "surname",
        "given_name",

        # identity / nationality
        "country",
        "country_code",
        "nationality",
        "nationality_code",

        # demographic / dates
        "sex",
        "gender",
        "date_of_birth",
        "date_of_issue",
        "date_of_expiry",
        "expiry_date",

        # extra OCR/MRZ data
        "registered_domicile",
        "issuing_authority",
        "raw_text",
        "ocr_raw_text",
        "image_quality_note",
        "detected_rotation_angle",
        "confidence_score",
        "status",
    ]

    for field in fields_from_fallback:
        if not final_result.get(field) and fallback_result.get(field):
            final_result[field] = fallback_result.get(field)

    # merge extra_data safely
    primary_extra = final_result.get("extra_data") or {}
    fallback_extra = fallback_result.get("extra_data") or {}

    if fallback_extra:
        merged_extra = dict(fallback_extra)
        merged_extra.update(primary_extra)
        final_result["extra_data"] = merged_extra

    return final_result
